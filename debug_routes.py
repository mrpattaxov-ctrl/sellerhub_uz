"""Debug routes for Uzum Seller Hub – extracted as a Flask Blueprint."""
from __future__ import annotations

import json
from datetime import date, datetime, timedelta
from urllib.request import Request, urlopen
from urllib.error import HTTPError

from flask import Blueprint, jsonify, request
from flask_login import login_required
from sqlalchemy import select, func

from core.auth_helpers import admin_required
from models import Shop, ProductGroup, Variant

# Late-bound imports from app – imported lazily inside init_debug_routes()
# to avoid circular imports (app.py imports this module to register the blueprint).
SessionLocal = None
http_json = None
_json_response = None
_get_admin_token = None
_today_app_tz = None
fetch_finance_sales_map = None
fetch_warehouse_expenses = None
find_first_array = None


def init_debug_routes(app_module):
    """Bind references from the main app module. Called once after app is created."""
    global SessionLocal, http_json, _json_response
    global _get_admin_token, _today_app_tz, fetch_finance_sales_map
    global fetch_warehouse_expenses, find_first_array
    SessionLocal = app_module.SessionLocal
    http_json = app_module.http_json
    _json_response = app_module._json_response
    _get_admin_token = app_module._get_admin_token
    _today_app_tz = app_module._today_app_tz
    fetch_finance_sales_map = app_module.fetch_finance_sales_map
    fetch_warehouse_expenses = app_module.fetch_warehouse_expenses
    find_first_array = app_module.find_first_array

debug_bp = Blueprint("debug_bp", __name__, url_prefix="/debug")


# ---------------------------------------------------------------------------
# Helper used only by debug routes
# ---------------------------------------------------------------------------
def _extract_page_items(data, page_size=100):
    """Extract items list from any paginated API response structure."""
    if isinstance(data, list):
        return data
    if not isinstance(data, dict):
        return []
    # Try nested structures: root, payload, data, result
    for key in [None, "payload", "data", "result"]:
        obj = data if key is None else data.get(key)
        if isinstance(obj, list):
            return obj
        if not isinstance(obj, dict):
            continue
        for items_key in ["content", "items", "data", "invoices", "list"]:
            val = obj.get(items_key)
            if isinstance(val, list):
                return val
    # Last resort: find any list of dicts
    for v in data.values():
        if isinstance(v, list) and v and isinstance(v[0], dict):
            return v
    return []


# ===================================================================
# 1. /debug/expenses
# ===================================================================
@debug_bp.get("/expenses")
@login_required
@admin_required
def debug_expenses():
    """Show raw expenses API response with detailed filtering info."""
    import datetime as _dt
    api_key = _get_admin_token()
    if not api_key:
        return jsonify({"error": "no token"}), 401
    with SessionLocal() as db:
        shops = db.execute(select(Shop)).scalars().all()
        shop_ids = [s.uzum_id for s in shops if s.uzum_id]
    ids_str = ",".join(str(s) for s in shop_ids)
    auth_val = (api_key if api_key.startswith("Bearer ") else f"Bearer {api_key}")
    headers = {
        "Authorization": auth_val,
        "Origin": "https://seller.uzum.uz",
        "Referer": "https://seller.uzum.uz/",
    }

    # Fetch using same source filters as main function
    from urllib.parse import quote as _quote
    all_payments = []
    for source_name in ["Склад", "Ombor"]:
        page = 0
        while True:
            url = (f"https://api.uzum.uz/api/seller/finance/expenses"
                   f"?page={page}&size=100"
                   f"&sources={_quote(source_name)}"
                   f"&shopIds={ids_str}")
            try:
                data = http_json(url, headers=headers)
            except Exception as e:
                break
            payments = []
            tp = 1
            if isinstance(data, dict):
                payload = data.get("payload")
                if isinstance(payload, dict):
                    payments = payload.get("payments") or []
                    tp = payload.get("totalPages", 1)
            all_payments.extend(payments)
            page += 1
            if page >= tp:
                break

    # Deduplicate
    seen_ids = set()
    unique_payments = []
    for p in all_payments:
        if not isinstance(p, dict):
            continue
        pid = p.get("id")
        if pid and pid in seen_ids:
            continue
        if pid:
            seen_ids.add(pid)
        unique_payments.append(p)

    _today_start = _dt.datetime.combine(_today_app_tz(), _dt.time(0, 0, 0))
    _today_end = _today_start + _dt.timedelta(days=1)
    _ts_start = int(_today_start.timestamp() * 1000)
    _ts_end = int(_today_end.timestamp() * 1000)

    debug_items = []
    today_items = []
    for p in unique_payments:
        dc = p.get("dateCreated") or 0
        ds = p.get("dateService") or 0
        dc_human = _dt.datetime.fromtimestamp(dc / 1000).strftime("%Y-%m-%d %H:%M") if dc else "N/A"
        ds_human = _dt.datetime.fromtimestamp(ds / 1000).strftime("%Y-%m-%d %H:%M") if ds else "N/A"
        is_today = _ts_start <= dc < _ts_end if dc else False
        item = {
            "name": p.get("name"),
            "code": p.get("code"),
            "source": p.get("source", ""),
            "paymentPrice": p.get("paymentPrice"),
            "dateCreated_human": dc_human,
            "dateService_human": ds_human,
            "is_today": is_today,
            "shopId": p.get("shopId"),
        }
        debug_items.append(item)
        if is_today:
            today_items.append(item)

    return jsonify({
        "total_warehouse_payments": len(unique_payments),
        "today_count": len(today_items),
        "today_range": f"{_today_start.strftime('%Y-%m-%d %H:%M')} to {_today_end.strftime('%Y-%m-%d %H:%M')}",
        "today_payments": today_items,
        "all_warehouse_payments": debug_items,
        "parsed_result": fetch_warehouse_expenses(shop_ids, api_key=api_key),
    })


# ===================================================================
# 2. /debug/finance
# ===================================================================
@debug_bp.get("/finance")
@login_required
@admin_required
def debug_finance():
    """
    Debugs the finance/sales sync for a shop.
    Shows raw API response, extracted identifiers, final sales map,
    and how many DB variants matched.

    Visit: /debug/finance?shop_id=YOUR_UZUM_SHOP_ID
    Optional: &days=30  (default 30)
    Optional: &page=0   (which page of raw orders to show, default 0)
    """
    shop_id = request.args.get("shop_id", "").strip()
    if not shop_id:
        return jsonify({"error": "Pass ?shop_id=YOUR_UZUM_SHOP_ID"})

    token = _get_admin_token()
    if not token:
        return jsonify({"error": "No API token configured"})

    days      = int(request.args.get("days", 30))
    raw_page  = int(request.args.get("page", 0))

    now_ts  = int(datetime.now().timestamp())
    past_ts = int((datetime.now() - timedelta(days=days)).timestamp())

    auth = f"Bearer {token}" if not token.startswith("Bearer ") else token
    headers = {"Authorization": auth, "Origin": "https://seller.uzum.uz", "Referer": "https://seller.uzum.uz/"}

    result = {
        "shop_id": shop_id,
        "date_range_days": days,
        "date_from_ts": past_ts,
        "date_to_ts": now_ts,
    }

    # -- 1. Raw API response (one page) --
    raw_response = {}
    for param in ("shopIds", "shopId"):
        url = (f"https://api-seller.uzum.uz/api/seller/finance/orders"
               f"?{param}={shop_id}&dateFrom={past_ts}&dateTo={now_ts}"
               f"&group=true&page={raw_page}&size=10")
        try:
            resp = http_json(url, headers=headers)
            items = find_first_array(resp, ["payload", "result", "content", "data",
                                            "items", "rows", "list", "orders", "orderItems"]) or []
            raw_response[param] = {
                "url": url,
                "items_on_page": len(items),
                "top_level_keys": list(resp.keys()) if isinstance(resp, dict) else [],
                "raw_first_item": items[0] if items else None,
                "raw_all_items_this_page": items,
            }
        except Exception as e:
            raw_response[param] = {"url": url, "error": str(e)}

    result["raw_api_response"] = raw_response

    # -- 2. Run the full finance fetch and show resulting sales_map --
    sales_map = None
    sales_map_error = None
    try:
        sales_map = fetch_finance_sales_map(shop_id, api_key=token, days=days)
    except Exception as e:
        sales_map_error = str(e)

    result["sales_map_error"] = sales_map_error
    result["sales_map_total_keys"] = len(sales_map) if sales_map else 0

    if sales_map:
        # Show entries with qty > 0, sorted by qty desc
        nonzero = sorted(
            [(k, v) for k, v in sales_map.items() if v.get("qty", 0) > 0],
            key=lambda x: -x[1].get("qty", 0),
        )
        result["sales_map_nonzero_count"] = len(nonzero)
        result["sales_map_top50"] = [
            {"identifier": k, "qty": v.get("qty"), "price": v.get("price")}
            for k, v in nonzero[:50]
        ]
        result["sales_map_sample_keys"] = list(sales_map.keys())[:20]
    else:
        result["sales_map_nonzero_count"] = 0
        result["sales_map_top50"] = []
        result["sales_map_sample_keys"] = []

    # -- 3. Compare sales_map against DB variants for this shop --
    db_match = {"matched": 0, "unmatched": 0, "matched_variants": [], "unmatched_variants": []}
    try:
        with SessionLocal() as db:
            shop_obj = db.execute(
                select(Shop).where(Shop.uzum_id == shop_id)
            ).scalar_one_or_none()

            if shop_obj and sales_map:
                rows = db.execute(
                    select(Variant, ProductGroup.name.label("group_name"))
                    .join(ProductGroup, Variant.group_id == ProductGroup.id)
                    .where(ProductGroup.shop_id == shop_obj.id)
                    .where(ProductGroup.is_archived == False)
                ).all()

                for v, group_name in rows:
                    sku     = (v.sku or "").strip()
                    barcode = (v.barcode or "").strip()
                    sku_id  = (v.uzum_sku_id or "").strip()

                    data = (sales_map.get(sku) or sales_map.get(sku.upper()) or
                            (sales_map.get(barcode) or sales_map.get(barcode.upper()) if barcode else None) or
                            (sales_map.get(sku_id) if sku_id else None))

                    entry = {
                        "variant_id": v.id,
                        "sku": sku,
                        "barcode": barcode,
                        "uzum_sku_id": sku_id,
                        "group_name": group_name[:60],
                        "db_sales_30d": v.sales_30d_finance,
                    }

                    if data:
                        db_match["matched"] += 1
                        entry["api_qty"] = data.get("qty")
                        entry["matched_by"] = (
                            "sku" if sales_map.get(sku) or sales_map.get(sku.upper()) else
                            "barcode" if barcode and (sales_map.get(barcode) or sales_map.get(barcode.upper())) else
                            "uzum_sku_id"
                        )
                        if len(db_match["matched_variants"]) < 30:
                            db_match["matched_variants"].append(entry)
                    else:
                        db_match["unmatched"] += 1
                        if len(db_match["unmatched_variants"]) < 30:
                            db_match["unmatched_variants"].append(entry)
            elif not shop_obj:
                db_match["error"] = f"Shop {shop_id} not found in DB — sync products first"
    except Exception as e:
        db_match["error"] = str(e)

    result["db_variant_match"] = db_match

    return jsonify(result)


# ===================================================================
# 3. /debug/api-probe
# ===================================================================
@debug_bp.get("/api-probe")
@login_required
@admin_required
def debug_api_probe():
    """Probe Uzum API endpoints to find which ones work with the current Bearer token."""
    shop_id = request.args.get("shop_id", "")
    token = _get_admin_token()
    if not token:
        return jsonify({"error": "No token set. Go to Settings and save your Uzum credentials first."})

    auth_header = token if token.startswith("Bearer ") else f"Bearer {token}"
    headers = {
        "Authorization": auth_header,
        "Accept": "application/json",
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 Chrome/120.0.0.0 Safari/537.36",
        "Origin": "https://seller.uzum.uz",
        "Referer": "https://seller.uzum.uz/",
    }

    results = {}
    now_ts = int(datetime.now().timestamp() * 1000)
    past_ts = int((datetime.now() - timedelta(days=30)).timestamp() * 1000)

    product_id = request.args.get("product_id", "1054356")

    shop_id_probe = request.args.get("shop_id", "5983")

    from datetime import date as _date
    _today = _date.today().isoformat()
    _week_ago = (_date.today() - timedelta(days=7)).isoformat()

    import urllib.parse as _up
    _analytics_query = _up.quote(json.dumps({
        "timezone": "Asia/Tashkent",
        "measures": ["SellerReportProductFunnelSku.sum_views"],
        "dimensions": [],
        "timeDimensions": [{"dimension": "SellerReportProductFunnelSku.date", "dateRange": [_week_ago, _today], "granularity": "day"}],
        "filters": []
    }))

    candidates = [
        # Analytics API (Cube.js) -- test if token works
        f"https://analytics-seller.uzum.uz/cubejs-api/v1/load?query={_analytics_query}&queryType=multi",
        # Single product detail with description (user-suggested)
        f"https://api-seller.uzum.uz/api/seller/shop/{shop_id_probe}/product/{product_id}/description-response",
        # Single product via shop
        f"https://api-seller.uzum.uz/api/seller/shop/{shop_id_probe}/product?productId={product_id}",
        # SKU/stock bulk endpoints
        f"https://api-seller.uzum.uz/api/seller/shop/{shop_id_probe}/sku-list?page=0&size=5",
        # Product list with SKU data
        f"https://api-seller.uzum.uz/api/seller/product?page=0&size=1&withSku=true",
    ]

    for url in candidates:
        try:
            req = Request(url, headers=headers)
            with urlopen(req, timeout=10) as resp:
                body = json.loads(resp.read().decode())
                # Return full body for 200s so we can inspect structure
                results[url] = {"status": resp.status, "body": body}
        except HTTPError as e:
            results[url] = {"status": e.code, "error": e.read().decode(errors="ignore")[:300]}
        except Exception as e:
            results[url] = {"error": str(e)}

    return jsonify({"token_prefix": auth_header[:30] + "...", "results": results})


# ===================================================================
# 4. /debug/rate-limit-test
# ===================================================================
@debug_bp.get("/rate-limit-test")
@login_required
@admin_required
def debug_rate_limit_test():
    """Send bursts of requests to Uzum API to discover rate limits.
    Usage: /debug/rate-limit-test?shop_id=10945&rps=5&total=30
      rps   = requests per second (default 5, max 100)
      total = total requests to send (default 30, max 500)
    """
    import time as _time
    from concurrent.futures import ThreadPoolExecutor, as_completed

    shop_id = request.args.get("shop_id", "").strip()
    if not shop_id:
        with SessionLocal() as db:
            shop = db.execute(select(Shop)).scalars().first()
            if shop:
                shop_id = shop.uzum_id

    if not shop_id:
        return _json_response({"error": "shop_id required"}, 400)

    rps = min(100, max(1, int(request.args.get("rps", "5"))))
    total = min(500, max(1, int(request.args.get("total", "30"))))

    token = _get_admin_token()
    if not token:
        return _json_response({"error": "No Uzum token configured"}, 400)

    auth = token if token.startswith("Bearer ") else f"Bearer {token}"
    today_ts = int(datetime.combine(date.today(), datetime.min.time()).timestamp())
    yesterday_ts = int(datetime.combine(date.today() - timedelta(days=1), datetime.min.time()).timestamp())

    # Use a lightweight finance endpoint (1 day, page 0, small size)
    test_url = (
        f"https://api-seller.uzum.uz/api/seller/finance/orders"
        f"?shopIds={shop_id}&dateFrom={yesterday_ts}&dateTo={today_ts}"
        f"&group=true&page=0&size=10"
    )

    results = []

    def _send_one(req_num):
        headers = {
            "Authorization": auth,
            "Accept": "application/json",
            "User-Agent": "Mozilla/5.0",
            "Origin": "https://seller.uzum.uz",
            "Referer": "https://seller.uzum.uz/",
        }
        t0 = _time.perf_counter()
        try:
            req_obj = Request(test_url, headers=headers)
            with urlopen(req_obj, timeout=15) as resp:
                body = resp.read()
                elapsed = round((_time.perf_counter() - t0) * 1000)
                return {
                    "req": req_num,
                    "status": resp.status,
                    "ms": elapsed,
                    "size": len(body),
                }
        except HTTPError as e:
            elapsed = round((_time.perf_counter() - t0) * 1000)
            body_text = e.read().decode(errors="ignore")[:200]
            return {
                "req": req_num,
                "status": e.code,
                "ms": elapsed,
                "error": body_text,
            }
        except Exception as e:
            elapsed = round((_time.perf_counter() - t0) * 1000)
            return {
                "req": req_num,
                "status": 0,
                "ms": elapsed,
                "error": str(e)[:200],
            }

    # Send requests in bursts of `rps` per second
    start_time = _time.perf_counter()
    sent = 0
    burst_num = 0

    while sent < total:
        burst_start = _time.perf_counter()
        burst_size = min(rps, total - sent)
        burst_num += 1

        with ThreadPoolExecutor(max_workers=burst_size) as pool:
            futures = []
            for i in range(burst_size):
                futures.append(pool.submit(_send_one, sent + i + 1))
            for f in as_completed(futures):
                results.append(f.result())

        sent += burst_size

        # Wait until 1 second has passed since burst start
        elapsed_in_burst = _time.perf_counter() - burst_start
        if elapsed_in_burst < 1.0 and sent < total:
            _time.sleep(1.0 - elapsed_in_burst)

    total_time = round((_time.perf_counter() - start_time) * 1000)
    results.sort(key=lambda x: x["req"])

    # Summary
    statuses = {}
    latencies = []
    errors = []
    for r in results:
        s = r["status"]
        statuses[s] = statuses.get(s, 0) + 1
        latencies.append(r["ms"])
        if s != 200:
            errors.append(r)

    return _json_response({
        "test_config": {
            "shop_id": shop_id,
            "rps": rps,
            "total_requests": total,
            "url": test_url,
        },
        "summary": {
            "total_time_ms": total_time,
            "actual_rps": round(total / (total_time / 1000), 1) if total_time > 0 else 0,
            "status_counts": statuses,
            "avg_latency_ms": round(sum(latencies) / len(latencies)) if latencies else 0,
            "min_latency_ms": min(latencies) if latencies else 0,
            "max_latency_ms": max(latencies) if latencies else 0,
            "success_rate": f"{statuses.get(200, 0)}/{total}",
            "rate_limited": statuses.get(429, 0) > 0,
        },
        "errors": errors[:20],
        "all_results": results,
    })


# ===================================================================
# 5. /debug/sync-trace
# ===================================================================
@debug_bp.get("/sync-trace")
@login_required
@admin_required
def debug_sync_trace():
    """
    Runs a mini-sync for one shop and returns a full trace report.
    Visit: /debug/sync-trace?shop_id=YOUR_SHOP_ID
    """
    shop_id = request.args.get("shop_id", "").strip()
    if not shop_id:
        return jsonify({"error": "Pass ?shop_id=YOUR_UZUM_SHOP_ID"})

    token = _get_admin_token()
    if not token:
        return jsonify({"error": "No API token configured"})

    auth = f"Bearer {token}" if not token.startswith("Bearer ") else token
    headers = {"Authorization": auth, "Origin": "https://seller.uzum.uz", "Referer": "https://seller.uzum.uz/"}

    report = {
        "shop_id": shop_id,
        "sku_list": {"pages": 0, "total_items": 0, "sample": []},
        "product_list": {"total": 0, "for_this_shop": 0, "statuses": {}, "sample": []},
        "db_after": {"active": 0, "archived": 0, "sort_order_zero": 0},
        "problems": [],
    }

    # Step 1: probe sku-list page 0
    try:
        raw = http_json(
            f"https://api-seller.uzum.uz/api/seller/shop/{shop_id}/sku-list?page=0&size=20",
            headers=headers
        )
        items = raw.get("skuList") or []
        report["sku_list"]["pages"] = raw.get("totalPages")
        report["sku_list"]["total_items"] = raw.get("totalElements") or raw.get("total") or len(items)
        for it in items[:5]:
            report["sku_list"]["sample"].append({
                "productId": it.get("productId"),
                "skuTitle": it.get("skuTitle"),
                "productTitle": it.get("productTitle"),
                "quantityActive": it.get("quantityActive"),
            })
    except Exception as e:
        report["problems"].append(f"sku-list error: {e}")

    # Step 2: probe product list page 0
    try:
        raw2 = http_json(
            "https://api-seller.uzum.uz/api/seller/product?page=0&size=100",
            headers=headers
        )
        payload = raw2.get("payload") or {}
        plist = payload.get("products") or []
        report["product_list"]["total"] = payload.get("totalProductAmount", 0)
        for p in plist:
            sid = str(p.get("shopId") or "")
            sv = (p.get("status") or {}).get("value") or "UNKNOWN"
            report["product_list"]["statuses"][sv] = report["product_list"]["statuses"].get(sv, 0) + 1
            if sid == shop_id:
                report["product_list"]["for_this_shop"] += 1
                if len(report["product_list"]["sample"]) < 5:
                    report["product_list"]["sample"].append({
                        "id": p.get("id"),
                        "title": (p.get("title") or "")[:50],
                        "shopId": sid,
                        "status": sv,
                    })
    except Exception as e:
        report["problems"].append(f"product-list error: {e}")

    # Step 3: check current DB state for this shop
    try:
        with SessionLocal() as db:
            shop_obj = db.execute(select(Shop).where(Shop.uzum_id == shop_id)).scalar_one_or_none()
            if shop_obj:
                rows = db.execute(
                    select(
                        ProductGroup.id,
                        ProductGroup.name,
                        ProductGroup.is_archived,
                        ProductGroup.uzum_sort_order,
                        ProductGroup.uzum_product_id,
                    ).where(ProductGroup.shop_id == shop_obj.id)
                ).all()
                report["db_after"]["active"] = sum(1 for r in rows if not r.is_archived)
                report["db_after"]["archived"] = sum(1 for r in rows if r.is_archived)
                report["db_after"]["sort_order_zero"] = sum(1 for r in rows if r.uzum_sort_order == 0)
                report["db_after"]["total"] = len(rows)
                # Flag any suspicious records: not archived but sort_order=0, or archived but sort_order>0
                for r in rows:
                    if not r.is_archived and r.uzum_sort_order == 0:
                        report["problems"].append(
                            f"ACTIVE but sort_order=0: id={r.id} name='{(r.name or '')[:40]}' uzum_pid={r.uzum_product_id}"
                        )
                    if r.is_archived and r.uzum_sort_order > 0:
                        report["problems"].append(
                            f"ARCHIVED but sort_order>0: id={r.id} name='{(r.name or '')[:40]}' sort={r.uzum_sort_order}"
                        )
            else:
                report["problems"].append(f"Shop {shop_id} not found in DB")
    except Exception as e:
        report["problems"].append(f"DB check error: {e}")

    return jsonify(report)


# ===================================================================
# 6. /debug/all-shops
# ===================================================================
@debug_bp.get("/all-shops")
@login_required
@admin_required
def debug_all_shops():
    """
    Pages through ALL products from /api/seller/product and reports
    every shop found, how many products each has, and their statuses.
    Also compares against what is stored in the local DB.
    Visit: /debug/all-shops
    """
    token = _get_admin_token()
    if not token:
        return jsonify({"error": "No API token configured"})

    auth = f"Bearer {token}" if not token.startswith("Bearer ") else token
    headers = {"Authorization": auth, "Origin": "https://seller.uzum.uz", "Referer": "https://seller.uzum.uz/"}

    # Page through ALL products
    shops: dict = {}   # shopId -> {count, statuses, sample}
    page = 0
    total_products = None
    collected = 0
    errors = []

    while True:
        try:
            raw = http_json(
                f"https://api-seller.uzum.uz/api/seller/product?page={page}&size=100",
                headers=headers,
            )
        except Exception as e:
            errors.append(f"page {page}: {e}")
            break

        payload = raw.get("payload") or {}
        plist   = payload.get("products") or []

        if total_products is None:
            total_products = int(payload.get("totalProductAmount") or 0)

        for p in plist:
            sid    = str(p.get("shopId") or "0")
            status = (p.get("status") or {}).get("value") or "UNKNOWN"

            if sid not in shops:
                shops[sid] = {"product_count": 0, "statuses": {}, "sample_products": []}

            shops[sid]["product_count"] += 1
            shops[sid]["statuses"][status] = shops[sid]["statuses"].get(status, 0) + 1

            if len(shops[sid]["sample_products"]) < 3:
                shops[sid]["sample_products"].append({
                    "id": p.get("id"),
                    "title": (p.get("title") or "")[:60],
                    "status": status,
                })

        collected += len(plist)
        if not plist or (total_products and collected >= total_products):
            break
        if page >= 200:
            errors.append("stopped at page 200 safety limit")
            break
        page += 1

    # Compare with DB
    db_shops = []
    try:
        with SessionLocal() as db:
            shop_rows = db.execute(select(Shop)).scalars().all()
            for s in shop_rows:
                group_count = db.execute(
                    select(func.count()).where(ProductGroup.shop_id == s.id)
                ).scalar()
                active_count = db.execute(
                    select(func.count()).where(
                        ProductGroup.shop_id == s.id,
                        ProductGroup.is_archived == False,
                    )
                ).scalar()
                db_shops.append({
                    "db_id": s.id,
                    "uzum_id": s.uzum_id,
                    "name": s.name,
                    "total_groups": group_count,
                    "active_groups": active_count,
                    "archived_groups": group_count - active_count,
                    "in_api": s.uzum_id in shops,
                })
    except Exception as e:
        errors.append(f"DB check error: {e}")

    # Flag shops in API but not in DB
    missing_from_db = [sid for sid in shops if sid not in [s["uzum_id"] for s in db_shops]]

    return jsonify({
        "total_products_api": total_products,
        "pages_fetched": page + 1,
        "shops_found_in_api": shops,
        "shops_in_db": db_shops,
        "shops_in_api_but_not_in_db": missing_from_db,
        "errors": errors,
    })


# ===================================================================
# 7. /debug/shop-products
# ===================================================================
@debug_bp.get("/shop-products")
@login_required
@admin_required
def debug_shop_products():
    """
    Fetches ALL products for a single shop from the Uzum API (all pages),
    shows them in both raw API order and sorted order (newest ID first),
    and compares against what is in the local DB.
    Visit: /debug/shop-products?shop_id=YOUR_UZUM_SHOP_ID
    """
    shop_id = request.args.get("shop_id", "").strip()
    if not shop_id:
        return jsonify({"error": "Pass ?shop_id=YOUR_UZUM_SHOP_ID"})

    token = _get_admin_token()
    if not token:
        return jsonify({"error": "No API token configured"})

    auth = f"Bearer {token}" if not token.startswith("Bearer ") else token
    headers = {"Authorization": auth, "Origin": "https://seller.uzum.uz", "Referer": "https://seller.uzum.uz/"}

    raw_products = []   # raw API order
    statuses: dict = {}
    page = 0
    total_products = None
    collected = 0
    errors = []

    # Fetch all pages
    while True:
        try:
            raw = http_json(
                f"https://api-seller.uzum.uz/api/seller/product?page={page}&size=100",
                headers=headers,
            )
        except Exception as e:
            errors.append(f"page {page}: {e}")
            break

        payload = raw.get("payload") or {}
        plist   = payload.get("products") or []

        if total_products is None:
            total_products = int(payload.get("totalProductAmount") or 0)

        for p in plist:
            if str(p.get("shopId") or "") != shop_id:
                continue
            status = (p.get("status") or {}).get("value") or "UNKNOWN"
            statuses[status] = statuses.get(status, 0) + 1
            raw_products.append(p)

        collected += len(plist)
        if not plist or (total_products and collected >= total_products):
            break
        if page >= 200:
            errors.append("stopped at page 200 safety limit")
            break
        page += 1

    # Sorted order: newest product ID first (same order sync writes to DB)
    sorted_products = sorted(raw_products, key=lambda p: -int(p.get("id") or 0))

    def fmt(lst):
        return [
            {
                "sync_position": i + 1,
                "id": p.get("id"),
                "title": (p.get("title") or "")[:70],
                "status": (p.get("status") or {}).get("value") or "UNKNOWN",
            }
            for i, p in enumerate(lst)
        ]

    # DB comparison
    db_products = []
    try:
        with SessionLocal() as db:
            shop_obj = db.execute(
                select(Shop).where(Shop.uzum_id == shop_id)
            ).scalar_one_or_none()
            if shop_obj:
                rows = db.execute(
                    select(ProductGroup)
                    .where(ProductGroup.shop_id == shop_obj.id)
                    .order_by(ProductGroup.uzum_sort_order.asc())
                ).scalars().all()
                db_products = [
                    {
                        "db_position": r.uzum_sort_order,
                        "id": r.uzum_product_id,
                        "title": (r.name or "")[:70],
                        "is_archived": r.is_archived,
                    }
                    for r in rows
                ]
    except Exception as e:
        errors.append(f"DB error: {e}")

    return jsonify({
        "shop_id": shop_id,
        "total_products_in_api_all_shops": total_products,
        "pages_fetched": page + 1,
        "products_found_for_this_shop": len(raw_products),
        "status_breakdown": statuses,
        "note": "api_raw_order = random Uzum internal order. sorted_order = what sync writes to DB (newest ID first).",
        "api_raw_order": fmt(raw_products),
        "sorted_order_sync_will_use": fmt(sorted_products),
        "db_current_order": db_products,
        "errors": errors,
    })


# ===================================================================
# 8. /debug/raw-api
# ===================================================================
@debug_bp.get("/raw-api")
@login_required
@admin_required
def debug_raw_api():
    """
    Returns the raw JSON from Uzum API for inspection.
    Visit: /debug/raw-api?shop_id=YOUR_UZUM_SHOP_ID

    Query params:
      shop_id  - required, Uzum shop ID
      page     - page number (default 0)
      size     - items per page (default 5)
      source   - "products" or "skulist" (default shows both)
    """
    shop_id = request.args.get("shop_id", "").strip()
    if not shop_id:
        return jsonify({"error": "Pass ?shop_id=YOUR_UZUM_SHOP_ID"})

    token = _get_admin_token()
    if not token:
        return jsonify({"error": "No API token configured"})

    auth = f"Bearer {token}" if not token.startswith("Bearer ") else token
    headers = {"Authorization": auth, "Origin": "https://seller.uzum.uz", "Referer": "https://seller.uzum.uz/"}

    page   = int(request.args.get("page", 0))
    size   = int(request.args.get("size", 5))
    source = request.args.get("source", "both").lower()

    result = {"shop_id": shop_id, "page": page, "size": size}

    if source in ("products", "both"):
        try:
            raw = http_json(
                f"https://api-seller.uzum.uz/api/seller/product?page={page}&size={size}",
                headers=headers,
            )
            # Show full raw response and pull out just this shop's products for easy reading
            payload   = raw.get("payload") or {}
            all_prods = payload.get("products") or []
            this_shop = [p for p in all_prods if str(p.get("shopId") or "") == shop_id]

            result["product_list"] = {
                "url": f"/api/seller/product?page={page}&size={size}",
                "totalProductAmount": payload.get("totalProductAmount"),
                "totalPages": payload.get("totalPages"),
                "products_on_this_page": len(all_prods),
                "products_for_shop": len(this_shop),
                "raw_first_product_ALL_fields": all_prods[0] if all_prods else None,
                "raw_first_product_this_shop": this_shop[0] if this_shop else None,
                "all_products_this_page": all_prods,
            }
        except Exception as e:
            result["product_list"] = {"error": str(e)}

    if source in ("skulist", "both"):
        try:
            raw = http_json(
                f"https://api-seller.uzum.uz/api/seller/shop/{shop_id}/sku-list?page={page}&size={size}",
                headers=headers,
            )
            items = raw.get("skuList") or []
            result["sku_list"] = {
                "url": f"/api/seller/shop/{shop_id}/sku-list?page={page}&size={size}",
                "totalPages": raw.get("totalPages"),
                "totalElements": raw.get("totalElements") or raw.get("total"),
                "items_on_this_page": len(items),
                "raw_first_sku_ALL_fields": items[0] if items else None,
                "all_items_this_page": items,
            }
        except Exception as e:
            result["sku_list"] = {"error": str(e)}

    return jsonify(result)


# ===================================================================
# 9. /debug/sku-list
# ===================================================================
@debug_bp.get("/sku-list")
@login_required
@admin_required
def debug_sku_list():
    """
    Probe alternative Uzum API endpoints to find the one that returns ALL products.
    Visit: /debug/sku-list?shop_id=YOUR_SHOP_ID
    """
    shop_id = request.args.get("shop_id", "").strip()
    if not shop_id:
        return jsonify({"error": "Pass ?shop_id=YOUR_UZUM_SHOP_ID"})

    token = _get_admin_token()
    if not token:
        return jsonify({"error": "No Uzum token set"})

    auth = token if token.startswith("Bearer ") else f"Bearer {token}"
    hdrs = {"Authorization": auth, "Origin": "https://seller.uzum.uz", "Referer": "https://seller.uzum.uz/"}

    def _probe(url):
        try:
            req = Request(url, headers=hdrs)
            with urlopen(req, timeout=12) as resp:
                body = json.loads(resp.read().decode())
                return {"http": resp.status, "body": body}
        except HTTPError as e:
            return {"http": e.code, "error": e.read().decode(errors="ignore")[:400]}
        except Exception as e:
            return {"error": str(e)}

    # -- Special: full first-product dump from /api/seller/product --
    product_ep_result = {}
    for ep_url in [
        f"https://api-seller.uzum.uz/api/seller/product?page=0&size=3&shopId={shop_id}&withSku=true",
        f"https://api-seller.uzum.uz/api/seller/product?page=0&size=3&withSku=true",
    ]:
        r = _probe(ep_url)
        body = r.get("body", {})
        if isinstance(body, dict) and "payload" in body:
            payload = body["payload"]
            products = payload.get("products") or []
            product_ep_result[ep_url] = {
                "http": r.get("http"),
                "totalProductAmount": payload.get("totalProductAmount"),
                "products_in_page": len(products),
                "first_product_FULL": products[0] if products else None,
                "second_product_keys": list(products[1].keys()) if len(products) > 1 else [],
            }
            break
        else:
            product_ep_result[ep_url] = r

    results = {}

    # -- 1. Current sku-list (baseline) --
    r = _probe(f"https://api-seller.uzum.uz/api/seller/shop/{shop_id}/sku-list?page=0&size=5")
    body = r.get("body", {})
    sku_list = body.get("skuList") or []
    first_item = sku_list[0] if sku_list else {}
    results["sku-list (current)"] = {
        "http": r.get("http"),
        "top_keys": list(body.keys()),
        "totalPages": body.get("totalPages"),
        "skuLimit": body.get("skuLimit"),
        "invoiceSkuLimit": body.get("invoiceSkuLimit"),
        "items_in_page": len(sku_list),
        "item_keys": list(first_item.keys()),
        "characteristics_sample": first_item.get("characteristics"),
        "productFields_sample": first_item.get("productFields"),
        "first_item_titles": {"productTitle": first_item.get("productTitle"),
                              "skuTitle": first_item.get("skuTitle")},
    }

    # -- 2. Alternative product endpoints --
    alt_endpoints = [
        # General product list (not shop-specific)
        f"https://api-seller.uzum.uz/api/seller/product?page=0&size=5&shopId={shop_id}",
        f"https://api-seller.uzum.uz/api/seller/product?page=0&size=5",
        f"https://api-seller.uzum.uz/api/seller/product?page=0&size=5&withSku=true",
        # Shop-specific product (not sku-list)
        f"https://api-seller.uzum.uz/api/seller/shop/{shop_id}/product?page=0&size=5",
        f"https://api-seller.uzum.uz/api/seller/shop/{shop_id}/product?page=0&size=5&withSku=true",
        # Product v2/v3
        f"https://api-seller.uzum.uz/api/seller/v2/product?page=0&size=5&shopId={shop_id}",
        f"https://api-seller.uzum.uz/api/seller/v2/shop/{shop_id}/product?page=0&size=5",
        # Offer/offer-list
        f"https://api-seller.uzum.uz/api/seller/shop/{shop_id}/offer?page=0&size=5",
        f"https://api-seller.uzum.uz/api/seller/offer?page=0&size=5&shopId={shop_id}",
        # Warehouse stock
        f"https://api-seller.uzum.uz/api/seller/shop/{shop_id}/warehouse?page=0&size=5",
        f"https://api-seller.uzum.uz/api/seller/shop/{shop_id}/stock?page=0&size=5",
    ]

    for url in alt_endpoints:
        r = _probe(url)
        body = r.get("body", {})
        if not isinstance(body, dict):
            results[url] = r
            continue

        info = {"http": r.get("http"), "top_keys": list(body.keys())}

        # Dig into "payload" if present (Uzum wraps many responses in payload)
        payload = body.get("payload")
        if isinstance(payload, dict):
            info["payload_keys"] = list(payload.keys())
            for pk, pv in payload.items():
                if isinstance(pv, list) and pv:
                    info[f"payload.{pk}_count"] = len(pv)
                    first = pv[0]
                    if isinstance(first, dict):
                        info[f"payload.{pk}_first_keys"] = list(first.keys())
                elif isinstance(pv, (int, float, str, bool)) and pk in (
                    "totalElements", "total", "totalCount", "totalPages", "page", "size"
                ):
                    info[f"payload.{pk}"] = pv
        elif isinstance(payload, list):
            info["payload_is_list"] = True
            info["payload_count"] = len(payload)
            if payload and isinstance(payload[0], dict):
                info["payload_first_keys"] = list(payload[0].keys())

        # Also scan top-level lists/pagination
        for k, v in body.items():
            if isinstance(v, list) and v:
                info[f"list_{k}_count"] = len(v)
                if isinstance(v[0], dict):
                    info[f"list_{k}_first_keys"] = list(v[0].keys())
            elif k in ("totalPages", "totalElements", "total", "totalCount"):
                info[k] = v

        results[url] = info

    return jsonify({"shop_id": shop_id, "product_endpoint": product_ep_result, "results": results})


# ===================================================================
# 10. /debug/invoices
# ===================================================================
@debug_bp.get("/invoices")
@login_required
def debug_invoices():
    """Fetch ALL accepted invoices (all pages) for a shop and return raw JSON for debugging."""
    shop_id = request.args.get("shop_id", "").strip()
    status = request.args.get("status", "ACCEPTED").strip()
    page_size = 100  # max per page to reduce round-trips

    if not shop_id:
        # Use first shop from DB
        with SessionLocal() as s:
            shop = s.query(Shop).first()
            if shop:
                shop_id = shop.uzum_id

    if not shop_id:
        return _json_response({"error": "No shop_id provided and no shops in DB"}, 400)

    all_items = []
    page = 0

    while True:
        url = (
            f"https://api-seller.uzum.uz/api/seller/shop/{shop_id}/invoice"
            f"?page={page}&size={page_size}&statuses={status}"
        )
        try:
            data = http_json(url)
        except Exception as e:
            return _json_response({"error": f"API error on page {page}: {str(e)}"}, 500)

        items = _extract_page_items(data, page_size)
        if isinstance(items, list):
            all_items.extend(items)

        if not items:
            break
        page += 1
        if len(items) < page_size:
            break

    # Build debug summary
    sample = all_items[0] if all_items else None
    return _json_response({
        "shop_id": shop_id,
        "status_filter": status,
        "total_invoices": len(all_items),
        "total_pages_fetched": page + 1 if isinstance(data, list) else page,
        "sample_invoice_keys": list(sample.keys()) if sample and isinstance(sample, dict) else None,
        "sample_invoice": sample,
        "all_invoices": all_items,
    })


# ===================================================================
# 11. /debug/invoice-summary
# ===================================================================
@debug_bp.get("/invoice-summary")
@login_required
def debug_invoice_summary():
    """Fetch ALL invoices (all statuses, all pages), then fetch products for each, and aggregate by skuTitle."""
    shop_id = request.args.get("shop_id", "").strip()
    status = request.args.get("status", "ACCEPTED").strip()

    if not shop_id:
        with SessionLocal() as s:
            shop = s.query(Shop).first()
            if shop:
                shop_id = shop.uzum_id

    if not shop_id:
        return _json_response({"error": "No shop_id provided"}, 400)

    # Step 1: Fetch all invoices to get IDs (all pages)
    all_invoices = []
    page = 0
    page_size = 100
    while True:
        url = (
            f"https://api-seller.uzum.uz/api/seller/shop/{shop_id}/invoice"
            f"?page={page}&size={page_size}&statuses={status}"
        )
        try:
            data = http_json(url)
        except Exception as e:
            break

        items = _extract_page_items(data, page_size)
        if isinstance(items, list):
            all_invoices.extend(items)

        if not items:
            break
        page += 1
        if len(items) < page_size:
            break

    # Extract invoice IDs
    invoice_ids = []
    for inv in all_invoices:
        if isinstance(inv, dict) and inv.get("id"):
            invoice_ids.append(inv["id"])

    # Step 2: Fetch products for each invoice IN PARALLEL (10 concurrent)
    from concurrent.futures import ThreadPoolExecutor, as_completed

    summary = {}  # skuTitle -> {quantityAccepted, purchasePrice, invoices}
    errors = []

    def _fetch_invoice_products(inv_id):
        url = f"https://api-seller.uzum.uz/api/seller/shop/{shop_id}/invoice/getInvoiceProducts?invoiceId={inv_id}"
        try:
            return inv_id, http_json(url), None
        except Exception as e:
            return inv_id, None, str(e)

    with ThreadPoolExecutor(max_workers=10) as pool:
        futures = {pool.submit(_fetch_invoice_products, iid): iid for iid in invoice_ids}
        for future in as_completed(futures):
            inv_id, prod_data, err = future.result()
            if err:
                errors.append({"invoice_id": inv_id, "error": err})
                continue

            products = []
            if isinstance(prod_data, list):
                products = prod_data
            elif isinstance(prod_data, dict):
                products = prod_data.get("content") or prod_data.get("items") or prod_data.get("products") or prod_data.get("data") or []

            for p in products:
                if not isinstance(p, dict):
                    continue
                sku_list = p.get("skuForInvoiceDtoList") or []
                for sku in sku_list:
                    if not isinstance(sku, dict):
                        continue
                    sku_title = sku.get("skuTitle") or "Unknown"
                    qty_accepted = sku.get("quantityAccepted") or 0
                    qty_to_stock = sku.get("quantityToStock") or 0
                    price = sku.get("purchasePrice") or 0

                    if sku_title not in summary:
                        summary[sku_title] = {
                            "skuTitle": sku_title,
                            "totalQuantityAccepted": 0,
                            "totalQuantityToStock": 0,
                            "purchasePrice": price,
                            "invoiceCount": 0,
                            "totalCost": 0,
                        }
                    summary[sku_title]["totalQuantityAccepted"] += qty_accepted
                    summary[sku_title]["totalQuantityToStock"] += qty_to_stock
                    summary[sku_title]["invoiceCount"] += 1
                    summary[sku_title]["totalCost"] += qty_accepted * price

    # Sort by total quantity descending
    sorted_summary = sorted(summary.values(), key=lambda x: x["skuTitle"])

    grand_total_qty_accepted = sum(item["totalQuantityAccepted"] for item in sorted_summary)
    grand_total_qty_to_stock = sum(item["totalQuantityToStock"] for item in sorted_summary)
    grand_total_cost = sum(item["totalCost"] for item in sorted_summary)

    return _json_response({
        "shop_id": shop_id,
        "status_filter": status,
        "total_invoices": len(invoice_ids),
        "invoice_ids": invoice_ids,
        "unique_products": len(sorted_summary),
        "grand_total_quantity_accepted": grand_total_qty_accepted,
        "grand_total_quantity_to_stock": grand_total_qty_to_stock,
        "grand_total_cost": grand_total_cost,
        "errors": errors,
        "products_summary": sorted_summary,
    })


# ===================================================================
# 12. /debug/products-summary
# ===================================================================
@debug_bp.get("/products-summary")
@login_required
def debug_products_summary():
    """Fetch all product IDs from sku-list, then get SKU details for each product."""
    shop_id = request.args.get("shop_id", "").strip()

    if not shop_id:
        with SessionLocal() as s:
            shop = s.query(Shop).first()
            if shop:
                shop_id = shop.uzum_id

    if not shop_id:
        return _json_response({"error": "No shop_id provided. Usage: /debug/products-summary?shop_id=10945"}, 400)

    # Step 1: Fetch all SKU list pages to collect unique product IDs
    product_ids = set()
    page = 0
    page_size = 100
    total_pages = None

    while True:
        url = f"https://api-seller.uzum.uz/api/seller/shop/{shop_id}/sku-list?page={page}&size={page_size}"
        try:
            raw = http_json(url)
        except Exception as e:
            return _json_response({"error": f"sku-list page {page} error: {str(e)}"}, 500)

        items = raw.get("skuList") or []
        for item in items:
            pid = item.get("productId")
            if pid:
                product_ids.add(pid)

        if total_pages is None:
            total_pages = raw.get("totalPages") or 1

        page += 1
        if page >= total_pages:
            break

    # Step 2: Fetch SKU details for each product
    all_skus = []
    errors = []
    for pid in sorted(product_ids):
        url = f"https://api-seller.uzum.uz/api/seller/shop/{shop_id}/product?productId={pid}"
        try:
            data = http_json(url)
        except Exception as e:
            errors.append({"product_id": pid, "error": str(e)})
            continue

        if isinstance(data, dict) and "payload" in data:
            data = data["payload"]

        product_title = data.get("title") or data.get("productTitle") or "" if isinstance(data, dict) else ""
        skus = []
        if isinstance(data, dict):
            skus = data.get("skuList") or data.get("skus") or data.get("variants") or []

        for sku in skus:
            if not isinstance(sku, dict):
                continue
            all_skus.append({
                "productId": pid,
                "productTitle": product_title,
                "skuFullTitle": sku.get("skuFullTitle"),
                "quantityActive": sku.get("quantityActive") or 0,
                "quantityReturned": sku.get("quantityReturned") or 0,
                "quantityDefected": sku.get("quantityDefected") or 0,
                "quantitySold": sku.get("quantitySold") or 0,
                "price": sku.get("price") or 0,
            })

    # Totals
    total_active = sum(s["quantityActive"] for s in all_skus)
    total_returned = sum(s["quantityReturned"] for s in all_skus)
    total_defected = sum(s["quantityDefected"] for s in all_skus)
    total_sold = sum(s["quantitySold"] for s in all_skus)

    return _json_response({
        "shop_id": shop_id,
        "total_products": len(product_ids),
        "total_skus": len(all_skus),
        "total_quantityActive": total_active,
        "total_quantityReturned": total_returned,
        "total_quantityDefected": total_defected,
        "total_quantitySold": total_sold,
        "errors": errors,
        "skus": all_skus,
    })


# ===================================================================
# 13. /debug/lost-goods
# ===================================================================
@debug_bp.get("/lost-goods")
@login_required
def debug_lost_goods():
    """Calculate lost goods using 3 sources:
    - LEFT_OUT_REPORT CSV: SKU, "В продаже" (active), "Брак" (defected)
    - Invoice summary: totalQuantityAccepted per SKU
    - Products summary: quantitySold per SKU

    lost = totalQuantityAccepted - "В продаже" - "Брак" - quantitySold

    Usage: /debug/lost-goods?shop_id=5983
    """
    from concurrent.futures import ThreadPoolExecutor, as_completed
    import uuid, time as _time, csv, io

    shop_id = request.args.get("shop_id", "").strip()

    if not shop_id:
        with SessionLocal() as s:
            shop = s.query(Shop).first()
            if shop:
                shop_id = shop.uzum_id

    if not shop_id:
        return _json_response({"error": "No shop_id. Usage: /debug/lost-goods?shop_id=5983"}, 400)

    errors = []
    page_size = 100

    # -- Source 1: LEFT_OUT_REPORT -> active and defected per SKU --
    report_data = {}  # sku -> {active, defected}

    # Create the report
    create_url = "https://api-seller.uzum.uz/api/seller/documents/create"
    body = {
        "idempotencyKey": str(uuid.uuid4()),
        "jobType": "LEFT_OUT_REPORT",
        "contentType": "XLSX",
        "params": {"shopIds": [str(shop_id)]},
    }
    request_id = None
    try:
        create_resp = http_json(create_url, method="POST", body=body)
        payload = create_resp.get("payload", {}) if isinstance(create_resp, dict) else {}
        request_id = payload.get("requestId")
    except Exception as e:
        errors.append({"step": "report_create", "error": str(e)})

    # Poll until completed
    if request_id:
        status = "CREATED"
        for _ in range(40):
            _time.sleep(3)
            try:
                direct = http_json(f"https://api-seller.uzum.uz/api/seller/documents/v2/{request_id}")
                if isinstance(direct, dict):
                    p = direct.get("payload", direct)
                    if isinstance(p, dict) and p.get("status"):
                        status = p["status"]
                    elif p.get("link"):
                        status = "COMPLETED"
                    if not isinstance(p, dict) or "status" not in p:
                        status = "COMPLETED"
                if status == "COMPLETED":
                    break
            except Exception:
                pass
            # Also check list
            try:
                docs = http_json(
                    f"https://api-seller.uzum.uz/api/seller/documents/v2"
                    f"?jobFilters=LEFT_OUT_REPORT&shopIds={shop_id}&page=0&size=5"
                )
                items = []
                if isinstance(docs, dict):
                    p2 = docs.get("payload", docs)
                    if isinstance(p2, list):
                        items = p2
                    elif isinstance(p2, dict):
                        for k in ["content", "items", "data"]:
                            v = p2.get(k)
                            if isinstance(v, list):
                                items = v
                                break
                for item in items:
                    if isinstance(item, dict) and item.get("requestId") == request_id:
                        status = item.get("status", "")
                        break
            except Exception:
                pass
            if status == "COMPLETED":
                break

        # Download CSV
        if status == "COMPLETED":
            try:
                meta = http_json(f"https://api-seller.uzum.uz/api/seller/documents/v2/{request_id}")
                csv_link = None
                if isinstance(meta, dict):
                    p = meta.get("payload", meta)
                    if isinstance(p, dict):
                        csv_link = p.get("link") or p.get("fileLink") or p.get("downloadLink")
                    if not csv_link:
                        csv_link = meta.get("link") or meta.get("fileLink")
                if csv_link:
                    import urllib.request as _ur
                    req = _ur.Request(csv_link, headers={
                        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) Chrome/120.0.0.0",
                    })
                    with _ur.urlopen(req, timeout=30) as resp:
                        raw_bytes = resp.read()
                        text = ""
                        for enc in ["utf-8-sig", "utf-8", "cp1251", "latin-1"]:
                            try:
                                text = raw_bytes.decode(enc)
                                break
                            except Exception:
                                continue
                        first_line = text.split("\n")[0] if text else ""
                        delimiter = ";" if ";" in first_line else ","
                        reader = csv.DictReader(io.StringIO(text), delimiter=delimiter)
                        for row in reader:
                            # Find the SKU column and value columns
                            sku = row.get("SKU") or row.get("sku") or row.get("\u0410\u0440\u0442\u0438\u043a\u0443\u043b") or ""
                            sku = sku.strip()
                            if not sku:
                                continue
                            # active quantity
                            active = 0
                            for key in ["\u0412 \u043f\u0440\u043e\u0434\u0430\u0436\u0435", "\u0432 \u043f\u0440\u043e\u0434\u0430\u0436\u0435", "In sale", "Active"]:
                                if key in row:
                                    try:
                                        active = int(float(row[key].replace(",", ".").strip() or "0"))
                                    except (ValueError, AttributeError):
                                        pass
                                    break
                            # defected
                            defected = 0
                            for key in ["\u0411\u0440\u0430\u043a", "\u0431\u0440\u0430\u043a", "Defected", "Defect"]:
                                if key in row:
                                    try:
                                        defected = int(float(row[key].replace(",", ".").strip() or "0"))
                                    except (ValueError, AttributeError):
                                        pass
                                    break
                            report_data[sku] = {"active": active, "defected": defected}
            except Exception as e:
                errors.append({"step": "report_download", "error": str(e)})
        else:
            errors.append({"step": "report_poll", "error": f"Report not ready, status: {status}"})

    # -- Source 2: Invoice summary -> totalQuantityAccepted per SKU --
    invoice_totals = {}  # skuTitle -> totalAccepted
    all_invoices = []
    page = 0
    while True:
        url = (
            f"https://api-seller.uzum.uz/api/seller/shop/{shop_id}/invoice"
            f"?page={page}&size={page_size}&statuses=ACCEPTED"
        )
        try:
            data = http_json(url)
        except Exception:
            break
        items = _extract_page_items(data, page_size)
        if isinstance(items, list):
            all_invoices.extend(items)
        if not items or len(items) < page_size:
            break
        page += 1

    invoice_ids = [inv["id"] for inv in all_invoices if isinstance(inv, dict) and inv.get("id")]

    def _fetch_inv(inv_id):
        url = f"https://api-seller.uzum.uz/api/seller/shop/{shop_id}/invoice/getInvoiceProducts?invoiceId={inv_id}"
        try:
            return inv_id, http_json(url), None
        except Exception as e:
            return inv_id, None, str(e)

    with ThreadPoolExecutor(max_workers=10) as pool:
        futures = {pool.submit(_fetch_inv, iid): iid for iid in invoice_ids}
        for future in as_completed(futures):
            inv_id, prod_data, err = future.result()
            if err:
                errors.append({"step": "invoice", "id": inv_id, "error": err})
                continue
            products = []
            if isinstance(prod_data, list):
                products = prod_data
            elif isinstance(prod_data, dict):
                products = prod_data.get("content") or prod_data.get("items") or prod_data.get("products") or prod_data.get("data") or []
            for p in products:
                if not isinstance(p, dict):
                    continue
                for sku in (p.get("skuForInvoiceDtoList") or []):
                    if not isinstance(sku, dict):
                        continue
                    title = sku.get("skuTitle") or "Unknown"
                    qty = sku.get("quantityAccepted") or 0
                    invoice_totals[title] = invoice_totals.get(title, 0) + qty

    # -- Source 3: Products summary -> quantitySold per SKU --
    sold_data = {}  # skuFullTitle -> quantitySold
    product_ids = set()
    page = 0
    while True:
        url = f"https://api-seller.uzum.uz/api/seller/shop/{shop_id}/sku-list?page={page}&size={page_size}"
        try:
            raw = http_json(url)
        except Exception:
            break
        items = (raw.get("skuList") or []) if isinstance(raw, dict) else []
        for item in items:
            pid = item.get("productId")
            if pid:
                product_ids.add(pid)
        total_pages = (raw.get("totalPages") or 1) if isinstance(raw, dict) else 1
        page += 1
        if page >= total_pages:
            break

    def _fetch_prod(pid):
        url = f"https://api-seller.uzum.uz/api/seller/shop/{shop_id}/product?productId={pid}"
        try:
            return pid, http_json(url), None
        except Exception as e:
            return pid, None, str(e)

    with ThreadPoolExecutor(max_workers=10) as pool:
        futures = {pool.submit(_fetch_prod, pid): pid for pid in sorted(product_ids)}
        for future in as_completed(futures):
            pid, data, err = future.result()
            if err:
                errors.append({"step": "product", "id": pid, "error": err})
                continue
            if isinstance(data, dict) and "payload" in data:
                data = data["payload"]
            skus = []
            if isinstance(data, dict):
                skus = data.get("skuList") or data.get("skus") or data.get("variants") or []
            for sku in skus:
                if not isinstance(sku, dict):
                    continue
                qty_sold = sku.get("quantitySold") or 0
                # Only use skuFullTitle for matching (skuTitle is too short, e.g. "16-ЧЕРН")
                full_title = sku.get("skuFullTitle") or ""
                if full_title:
                    sold_data[full_title] = qty_sold

    # -- Calculate: lost = invoiced - active - defected - sold --
    # Only include SKUs that exist in report or sold (skip archived products)
    active_skus = sorted(set(
        list(report_data.keys()) +
        list(sold_data.keys())
    ))

    results = []
    t_invoiced = t_active = t_defected = t_sold = t_lost = 0

    for sku in active_skus:
        invoiced = invoice_totals.get(sku, 0)
        rd = report_data.get(sku, {})
        active = rd.get("active", 0)
        defected = rd.get("defected", 0)
        sold = sold_data.get(sku, 0)
        lost = invoiced - active - defected - sold

        t_invoiced += invoiced
        t_active += active
        t_defected += defected
        t_sold += sold
        t_lost += lost

        results.append({
            "sku": sku,
            "invoiced": invoiced,
            "active": active,
            "defected": defected,
            "sold": sold,
            "lost": lost,
            "_in_invoices": sku in invoice_totals,
            "_in_report": sku in report_data,
            "_in_sold": sku in sold_data,
        })

    show_all = request.args.get("all", "").lower() in ("true", "1", "yes")
    if not show_all:
        results = [r for r in results if r["lost"] != 0]

    # Debug: show sample keys from each source to diagnose mismatches
    sample_invoice_keys = sorted(invoice_totals.keys())[:5]
    sample_report_keys = sorted(report_data.keys())[:5]
    sample_sold_keys = sorted(sold_data.keys())[:5]

    return _json_response({
        "shop_id": shop_id,
        "formula": "lost = invoiced - active(\u0412 \u043f\u0440\u043e\u0434\u0430\u0436\u0435) - defected(\u0411\u0440\u0430\u043a) - sold",
        "total_invoices": len(invoice_ids),
        "total_products": len(product_ids),
        "report_skus": len(report_data),
        "sold_keys_count": len(sold_data),
        "total_skus": len(active_skus),
        "_debug_sample_invoice_keys": sample_invoice_keys,
        "_debug_sample_report_keys": sample_report_keys,
        "_debug_sample_sold_keys": sample_sold_keys,
        "totals": {
            "invoiced": t_invoiced,
            "active": t_active,
            "defected": t_defected,
            "sold": t_sold,
            "lost": t_lost,
        },
        "showing": "only lost/extra items (add ?all=true for all)" if not show_all else "all items",
        "errors": errors,
        "items": results,
    })


# ===================================================================
# 14. /debug/documents
# ===================================================================
@debug_bp.get("/documents")
@login_required
def debug_documents():
    """Fetch seller documents/reports. Usage: /debug/documents?shop_id=5983"""
    shop_id = request.args.get("shop_id", "").strip()
    page = request.args.get("page", "0").strip()
    size = request.args.get("size", "100").strip()
    jobs = request.args.get("jobs", "SELLS_REPORT,EXPENSES_REPORT,LEFT_OUT_REPORT,MARKED_SALES_REPORT,COMMISSIONER_REPORT,PAID_STORAGE_REPORT,SELLER_RETURN_PAID_STORAGE_REPORT,SELLER_STORAGE_REPORT,LEFT_OUT_REPORT_2024").strip()

    shop_ids_str = shop_id
    if not shop_ids_str:
        with SessionLocal() as s:
            shops = s.execute(select(Shop)).scalars().all()
            shop_ids_str = ",".join(sh.uzum_id for sh in shops if sh.uzum_id)

    url = (
        f"https://api-seller.uzum.uz/api/seller/documents/v2"
        f"?jobFilters={jobs}"
        f"&shopIds={shop_ids_str}"
        f"&page={page}&size={size}"
    )
    try:
        data = http_json(url)
    except Exception as e:
        return _json_response({"error": str(e), "url": url}, 500)

    return _json_response({
        "url": url,
        "raw_response": data,
    })


# ===================================================================
# 15. /debug/report
# ===================================================================
@debug_bp.get("/report")
@login_required
def debug_report():
    """Fetch, wait and download a report. Checks existing list first before creating.

    Usage: /debug/report?shop_id=5983&type=LEFT_OUT_REPORT
    Report types: SELLS_REPORT, EXPENSES_REPORT, LEFT_OUT_REPORT,
    MARKED_SALES_REPORT, COMMISSIONER_REPORT, PAID_STORAGE_REPORT,
    SELLER_RETURN_PAID_STORAGE_REPORT, SELLER_STORAGE_REPORT, LEFT_OUT_REPORT_2024
    """
    import uuid, json, time as _time, csv, io, urllib.request as _ur
    from datetime import date, datetime, timedelta

    shop_id     = request.args.get("shop_id", "").strip()
    report_type = request.args.get("type", "LEFT_OUT_REPORT").strip()
    # Date range — default to last 30 days
    today          = date.today()
    date_from      = request.args.get("date_from", (today - timedelta(days=30)).strftime("%Y-%m-%d")).strip()
    date_to        = request.args.get("date_to",   today.strftime("%Y-%m-%d")).strip()
    # Unix timestamps — same format Uzum uses in finance/orders API
    date_from_ts   = int(datetime.strptime(date_from, "%Y-%m-%d").timestamp())
    date_to_ts     = int(datetime.strptime(date_to,   "%Y-%m-%d").replace(hour=23, minute=59, second=59).timestamp())
    try:
        poll_interval_ms = max(100, int(request.args.get("poll_interval_ms", "3000")))
    except (ValueError, TypeError):
        poll_interval_ms = 3000
    poll_interval_s = poll_interval_ms / 1000.0

    t_total_start = _time.perf_counter()
    t_create_ms   = 0
    t_poll_ms     = 0
    t_download_ms = 0
    poll_attempts = 0

    # Reports that don't need a date range
    NO_DATE_TYPES = {"LEFT_OUT_REPORT", "LEFT_OUT_REPORT_2024", "SELLER_STORAGE_REPORT"}

    if not shop_id:
        with SessionLocal() as s:
            shop = s.query(Shop).first()
            if shop:
                shop_id = shop.uzum_id

    if not shop_id:
        return _json_response({"error": "No shop_id"}, 400)

    # ── helpers ──────────────────────────────────────────────────────────
    def _extract_items(data):
        if isinstance(data, list):
            return data
        if not isinstance(data, dict):
            return []
        p = data.get("payload", data)
        if isinstance(p, list):
            return p
        if isinstance(p, dict):
            for k in ["content", "items", "data", "list"]:
                v = p.get(k)
                if isinstance(v, list):
                    return v
        return []

    def _extract_link(meta):
        if not isinstance(meta, dict):
            return None
        p = meta.get("payload", meta)
        if isinstance(p, dict):
            lnk = p.get("link") or p.get("fileLink") or p.get("downloadLink")
            if lnk:
                return lnk
        return meta.get("link") or meta.get("fileLink") or meta.get("downloadLink")

    def _download_and_parse(file_url):
        req = _ur.Request(file_url, headers={
            "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) Chrome/120.0.0.0",
        })
        with _ur.urlopen(req, timeout=60) as resp:
            raw = resp.read()

        # XLSX/ZIP files start with PK magic bytes — parse directly, skip CSV attempt
        if raw[:2] == b'PK':
            try:
                import openpyxl as _xl, io as _io2
                wb = _xl.load_workbook(_io2.BytesIO(raw), read_only=True, data_only=True)
                ws = wb.active
                all_rows = list(ws.iter_rows(values_only=True))
                if not all_rows:
                    return [], [], "xlsx"
                headers = [str(c) if c is not None else "" for c in all_rows[0]]
                rows = [
                    dict(zip(headers, [str(c) if c is not None else "" for c in r]))
                    for r in all_rows[1:]
                ]
                return rows, headers, "xlsx"
            except Exception as e:
                raise RuntimeError(f"XLSX parse failed: {e}") from e

        # Try CSV parsing — normalize line endings first
        text = None
        for enc in ["utf-8-sig", "utf-8", "cp1251", "latin-1"]:
            try:
                text = raw.decode(enc)
                break
            except Exception:
                continue
        if text:
            text = text.replace('\r\n', '\n').replace('\r', '\n')
            first_line = text.split("\n")[0]
            delimiter = ";" if ";" in first_line else ","
            reader = csv.DictReader(io.StringIO(text), delimiter=delimiter)
            rows = list(reader)
            if rows or reader.fieldnames:
                return rows, reader.fieldnames or [], "csv"

        # Final XLSX fallback (in case magic bytes check missed it)
        try:
            import openpyxl as _xl2, io as _io3
            wb = _xl2.load_workbook(_io3.BytesIO(raw), read_only=True, data_only=True)
            ws = wb.active
            all_rows = list(ws.iter_rows(values_only=True))
            if not all_rows:
                return [], [], "xlsx"
            headers = [str(c) if c is not None else "" for c in all_rows[0]]
            rows = [dict(zip(headers, [str(c) if c is not None else "" for c in r])) for r in all_rows[1:]]
            return rows, headers, "xlsx"
        except Exception:
            pass
        return [], [], "unknown"

    # ── Step 1: Check existing list — only for NO_DATE types ─────────────
    # Date-based types (SELLS_REPORT etc.) must always create fresh so that
    # the requested date range is respected.  Reusing a cached report from a
    # different period would silently return wrong data.
    list_url = (
        f"https://api-seller.uzum.uz/api/seller/documents/v2"
        f"?jobFilters={report_type}&shopIds={shop_id}&page=0&size=20"
    )
    need_dates = report_type not in NO_DATE_TYPES
    request_id = None
    file_link = None
    source = None

    if not need_dates:
        # NO_DATE types: safe to reuse any completed report (no date mismatch risk)
        try:
            docs = http_json(list_url)
            items = _extract_items(docs)
            for item in items:
                if not isinstance(item, dict):
                    continue
                jt = item.get("jobType") or item.get("type") or ""
                st = item.get("status") or item.get("jobStatus") or ""
                if (not jt or jt == report_type) and st == "COMPLETED":
                    request_id = item.get("requestId") or item.get("id")
                    file_link = _extract_link(item) or item.get("link") or item.get("fileUrl")
                    source = "existing_list"
                    break
        except Exception:
            pass

    # ── Step 2: If not found, try to create ──────────────────────────────
    # Exact format captured from Uzum seller dashboard (browser network tab):
    #   POST /api/seller/documents/create
    #   {
    #     "idempotencyKey": "<uuid>",
    #     "jobType": "SELLS_REPORT",
    #     "contentType": "CSV",            ← TOP level, not inside params
    #     "params": {
    #       "returns": false,
    #       "group": false,                ← false, not true
    #       "dateFrom": 1776211200000,     ← milliseconds
    #       "dateTo":   1776297599000,
    #       "shopIds": [5983]
    #     }
    #   }
    create_error = None
    if not request_id:
        create_url = "https://api-seller.uzum.uz/api/seller/documents/create"
        create_headers = {
            "Origin":  "https://seller.uzum.uz",
            "Referer": "https://seller.uzum.uz/",
        }
        date_from_ms = date_from_ts * 1000
        date_to_ms   = date_to_ts   * 1000

        params = {
            "returns": False,
            "group":   False,
            "shopIds": [int(shop_id)],
        }
        if need_dates:
            params["dateFrom"] = date_from_ms
            params["dateTo"]   = date_to_ms

        body = {
            "idempotencyKey": str(uuid.uuid4()),
            "jobType":        report_type,
            "contentType":    "CSV",
            "params":         params,
        }

        _t_c0 = _time.perf_counter()
        try:
            create_resp = http_json(create_url, method="POST", body=body, headers=create_headers)
            t_create_ms = int((_time.perf_counter() - _t_c0) * 1000)
            p = create_resp.get("payload", {}) if isinstance(create_resp, dict) else {}
            request_id = p.get("requestId")
            if request_id:
                source = "created"
        except Exception as e:
            t_create_ms = int((_time.perf_counter() - _t_c0) * 1000)
            create_error = str(e)

        if not request_id:
            return _json_response({
                "error": f"Could not create report. Uzum error: {create_error}",
                "list_url": list_url,
                "body_sent": body,
            }, 500)

    # ── Step 3: Poll until COMPLETED ─────────────────────────────────────
    status = "COMPLETED" if file_link else "CREATED"
    last_raw = None

    _t_p0 = _time.perf_counter()
    if not file_link:
        for _ in range(40):
            _time.sleep(poll_interval_s)
            poll_attempts += 1
            try:
                direct = http_json(f"https://api-seller.uzum.uz/api/seller/documents/v2/{request_id}")
                last_raw = direct
                p = direct.get("payload", direct) if isinstance(direct, dict) else {}
                if isinstance(p, dict):
                    s = p.get("status") or p.get("jobStatus") or ""
                    if s:
                        status = s
                    lnk = _extract_link(direct)
                    if lnk:
                        file_link = lnk
                        status = "COMPLETED"
                if status == "COMPLETED":
                    break
            except Exception:
                pass

            # Also check the list
            try:
                docs2 = http_json(
                    f"https://api-seller.uzum.uz/api/seller/documents/v2"
                    f"?jobFilters={report_type}&shopIds={shop_id}&page=0&size=5"
                )
                for item in _extract_items(docs2):
                    if isinstance(item, dict) and item.get("requestId") == request_id:
                        status = item.get("status") or item.get("jobStatus") or status
                        lnk = _extract_link(item) or item.get("link") or item.get("fileUrl")
                        if lnk:
                            file_link = lnk
                        break
            except Exception:
                pass

            if status == "COMPLETED":
                break

    if status != "COMPLETED" or not file_link:
        # Try one last direct fetch for the link
        try:
            meta = http_json(f"https://api-seller.uzum.uz/api/seller/documents/v2/{request_id}")
            file_link = _extract_link(meta)
            last_raw = meta
        except Exception:
            pass

    t_poll_ms = int((_time.perf_counter() - _t_p0) * 1000)

    if not file_link:
        return _json_response({
            "error": f"Report not ready or no download link found. Status: {status}",
            "requestId": request_id,
            "source": source,
            "last_response": last_raw,
            "timings": {
                "create_ms": t_create_ms,
                "poll_ms": t_poll_ms,
                "poll_attempts": poll_attempts,
                "poll_interval_ms": poll_interval_ms,
                "poll_sleep_ms": poll_attempts * poll_interval_ms,
                "download_ms": 0,
                "total_ms": int((_time.perf_counter() - t_total_start) * 1000),
            },
        }, 408)

    # ── Step 4: Download and parse ────────────────────────────────────────
    # Cap rows sent to the browser so a multi-year report can't OOM the tab.
    # Full file is still reachable via csv_link.
    try:
        MAX_PREVIEW_ROWS = max(50, int(request.args.get("max_rows", "1000")))
    except (ValueError, TypeError):
        MAX_PREVIEW_ROWS = 1000
    _t_d0 = _time.perf_counter()
    try:
        rows, columns, fmt = _download_and_parse(file_link)
        t_download_ms = int((_time.perf_counter() - _t_d0) * 1000)
        total_rows = len(rows)
        preview_rows = rows[:MAX_PREVIEW_ROWS]
        truncated = total_rows > len(preview_rows)
        return _json_response({
            "shop_id": shop_id,
            "report_type": report_type,
            "source": source,
            "file_format": fmt,
            "requestId": request_id,
            "csv_link": file_link,
            "total_rows": total_rows,
            "preview_rows": len(preview_rows),
            "truncated": truncated,
            "max_preview_rows": MAX_PREVIEW_ROWS,
            "columns": columns,
            "data": preview_rows,
            "timings": {
                "create_ms": t_create_ms,
                "poll_ms": t_poll_ms,
                "poll_attempts": poll_attempts,
                "poll_interval_ms": poll_interval_ms,
                "poll_sleep_ms": poll_attempts * poll_interval_ms,
                "download_ms": t_download_ms,
                "total_ms": int((_time.perf_counter() - t_total_start) * 1000),
            },
        })
    except Exception as e:
        t_download_ms = int((_time.perf_counter() - _t_d0) * 1000)
        return _json_response({
            "error": f"File download/parse failed: {e}",
            "requestId": request_id,
            "csv_link": file_link,
            "timings": {
                "create_ms": t_create_ms,
                "poll_ms": t_poll_ms,
                "poll_attempts": poll_attempts,
                "poll_interval_ms": poll_interval_ms,
                "poll_sleep_ms": poll_attempts * poll_interval_ms,
                "download_ms": t_download_ms,
                "total_ms": int((_time.perf_counter() - t_total_start) * 1000),
            },
        }, 500)


# ===================================================================
# /debug/report-probe  — try multiple create body formats to find what works
# ===================================================================
@debug_bp.get("/report-probe")
@login_required
def debug_report_probe():
    """Probe the Uzum create-report endpoint with different body formats.
    Usage: /debug/report-probe?shop_id=5983&type=SELLS_REPORT
    """
    import uuid as _uuid
    shop_id = request.args.get("shop_id", "").strip()
    report_type = request.args.get("type", "SELLS_REPORT").strip()

    if not shop_id:
        with SessionLocal() as s:
            shop = s.query(Shop).first()
            if shop:
                shop_id = shop.uzum_id

    create_url = "https://api-seller.uzum.uz/api/seller/documents/create"

    import json as _json
    from datetime import date as _date, datetime as _datetime, timedelta as _td
    _NO_DATE = {"LEFT_OUT_REPORT", "LEFT_OUT_REPORT_2024", "SELLER_STORAGE_REPORT"}
    _today = _date.today()
    _dfrom_str = request.args.get("date_from", (_today - _td(days=30)).strftime("%Y-%m-%d"))
    _dto_str   = request.args.get("date_to",   _today.strftime("%Y-%m-%d"))
    # Convert YYYY-MM-DD to Unix timestamps (seconds) — same format Uzum finance API uses
    _dfrom_ts  = int(_datetime.strptime(_dfrom_str, "%Y-%m-%d").timestamp())
    _dto_ts    = int(_datetime.strptime(_dto_str,   "%Y-%m-%d").replace(hour=23, minute=59, second=59).timestamp())
    # Keep original strings too as fallback
    _dfrom = _dfrom_str
    _dto   = _dto_str

    # Verified working format (captured from seller.uzum.uz browser network tab):
    # - contentType at TOP level as "CSV" (NOT inside params, NOT "MS_EXCEL")
    # - group: false, returns: false INSIDE params (no language field)
    # - shopIds is an int array [int]
    # - dateFrom/dateTo are millisecond unix timestamps
    _has_dates = report_type not in _NO_DATE
    _dfrom_ms  = _dfrom_ts * 1000
    _dto_ms    = _dto_ts   * 1000

    def _params_good(with_dates=True):
        p = {"returns": False, "group": False, "shopIds": [int(shop_id)]}
        if with_dates and _has_dates:
            p["dateFrom"] = _dfrom_ms
            p["dateTo"]   = _dto_ms
        return p

    candidates = [
        # 1. The verified-working format — this should succeed
        {"idempotencyKey": str(_uuid.uuid4()), "jobType": report_type,
         "contentType": "CSV", "params": _params_good(True)},
        # 2. Same without dates (for NO_DATE types)
        {"idempotencyKey": str(_uuid.uuid4()), "jobType": report_type,
         "contentType": "CSV", "params": _params_good(False)},
        # 3. MS_EXCEL variant at top level (for comparison)
        {"idempotencyKey": str(_uuid.uuid4()), "jobType": report_type,
         "contentType": "MS_EXCEL", "params": _params_good(True)},
    ]

    _create_headers = {
        "Origin":  "https://seller.uzum.uz",
        "Referer": "https://seller.uzum.uz/",
    }
    results = []
    for body in candidates:
        try:
            resp = http_json(create_url, method="POST", body=body, headers=_create_headers)
            results.append({"body": body, "status": "OK", "response": resp})
            break  # stop at first success
        except Exception as e:
            results.append({"body": body, "status": "ERROR", "error": str(e)})

    return _json_response({
        "shop_id": shop_id,
        "report_type": report_type,
        "date_from": _dfrom,
        "date_to": _dto,
        "probe_results": results,
    })


# ===================================================================
# /debug/report-list  — dump raw list response to see document structure
# ===================================================================
@debug_bp.get("/report-list")
@login_required
def debug_report_list():
    """Dump the raw documents list to inspect structure.
    Usage: /debug/report-list?shop_id=5983&type=SELLS_REPORT
    """
    shop_id     = request.args.get("shop_id", "").strip()
    report_type = request.args.get("type", "SELLS_REPORT").strip()

    if not shop_id:
        with SessionLocal() as s:
            shop = s.query(Shop).first()
            if shop:
                shop_id = shop.uzum_id

    url = (
        f"https://api-seller.uzum.uz/api/seller/documents/v2"
        f"?jobFilters={report_type}&shopIds={shop_id}&page=0&size=5"
    )
    try:
        raw = http_json(url)
    except Exception as e:
        return _json_response({"error": str(e), "url": url}, 500)

    return _json_response({"url": url, "shop_id": shop_id, "report_type": report_type, "raw": raw})


# ===================================================================
# 16. /debug/raw-url
# ===================================================================
@debug_bp.get("/raw-url")
@login_required
def debug_raw_url():
    """Fetch any URL and return the raw response. Usage: /debug/raw-url?url=https://..."""
    url = request.args.get("url", "").strip()
    if not url:
        return _json_response({"error": "Pass ?url=https://..."}, 400)
    try:
        data = http_json(url)
    except Exception as e:
        return _json_response({"error": str(e)}, 500)
    return _json_response({
        "url": url,
        "response_type": type(data).__name__,
        "response_keys": list(data.keys()) if isinstance(data, dict) else None,
        "response": data,
    })


# ===================================================================
# 17. /debug/test-getproducts
# ===================================================================
@debug_bp.get("/test-getproducts")
@login_required
def debug_test_getproducts():
    """Test the getProducts endpoint. Usage: /debug/test-getproducts?shop_id=5983&size=500"""
    shop_id = request.args.get("shop_id", "").strip()
    size = request.args.get("size", "").strip()
    if not shop_id:
        return _json_response({"error": "Usage: /debug/test-getproducts?shop_id=5983&size=500  (omit size to test without it)"}, 400)
    base = f"https://api-seller.uzum.uz/api/seller/shop/{shop_id}/product/getProducts?filter=ALL&sortBy=id&order=descending"
    if size:
        url = f"{base}&page=0&size={size}"
    else:
        url = base
    try:
        data = http_json(url)
    except Exception as e:
        return _json_response({"error": str(e), "url": url}, 500)
    total = data.get("totalProductsAmount", "?")
    returned = len(data.get("productList", []))
    return _json_response({
        "url": url,
        "totalProductsAmount": total,
        "returnedCount": returned,
        "allLoaded": total == returned if isinstance(total, int) else "unknown",
        "firstProduct": data.get("productList", [None])[0] if data.get("productList") else None,
    })


# ===================================================================
# 18. /debug/sku-list-full
# ===================================================================
@debug_bp.get("/sku-list-full")
@login_required
def debug_sku_list_full():
    """Fetch return SKUs for a product. Usage: /debug/sku-list-full?shop_id=5983&product_id=311964"""
    shop_id = request.args.get("shop_id", "").strip()
    product_id = request.args.get("product_id", "").strip()

    if not shop_id:
        with SessionLocal() as s:
            shop = s.query(Shop).first()
            if shop:
                shop_id = shop.uzum_id

    if not shop_id or not product_id:
        return _json_response({"error": "Both shop_id and product_id are required. Usage: /debug/sku-list-full?shop_id=5983&product_id=311964"}, 400)

    url = f"https://api-seller.uzum.uz/api/seller/shop/{shop_id}/product?productId={product_id}"
    try:
        data = http_json(url)
    except Exception as e:
        return _json_response({"error": f"API error: {str(e)}"}, 500)

    return _json_response({
        "shop_id": shop_id,
        "product_id": product_id,
        "raw_response": data,
    })


# ===================================================================
# 19. /debug/returns
# ===================================================================
@debug_bp.get("/returns")
@login_required
def debug_returns():
    """Fetch ALL return invoices (all pages) for a shop and return raw JSON."""
    shop_id = request.args.get("shop_id", "").strip()
    statuses = request.args.get("statuses", "CREATED,SENT,IN_PROGRESS,MOVED_TO_DELIVERY,ASSEMBLED,COMPLETED,UTILIZED").strip()
    types = request.args.get("types", "FBS,DEFECTED,RETURN").strip()
    page_size = 100

    if not shop_id:
        with SessionLocal() as s:
            shop = s.query(Shop).first()
            if shop:
                shop_id = shop.uzum_id

    if not shop_id:
        return _json_response({"error": "No shop_id provided"}, 400)

    # First, get raw page 0 for debugging
    first_url = (
        f"https://api-seller.uzum.uz/api/seller/shop/{shop_id}/return"
        f"?page=0&size={page_size}&statuses={statuses}&types={types}"
    )
    try:
        raw_first_page = http_json(first_url)
    except Exception as e:
        return _json_response({"error": f"API error: {str(e)}", "url": first_url}, 500)

    all_items = []
    page = 0
    while True:
        url = (
            f"https://api-seller.uzum.uz/api/seller/shop/{shop_id}/return"
            f"?page={page}&size={page_size}&statuses={statuses}&types={types}"
        )
        try:
            data = http_json(url)
        except Exception as e:
            break

        # Try multiple response structures
        items = []
        total_pages = 1
        if isinstance(data, list):
            items = data
        elif isinstance(data, dict):
            payload = data.get("payload")
            if isinstance(payload, dict):
                items = payload.get("content") or payload.get("items") or payload.get("data") or []
                total_pages = payload.get("totalPages") or payload.get("total_pages") or 1
            elif isinstance(payload, list):
                items = payload
            else:
                items = data.get("content") or data.get("items") or data.get("data") or []
                total_pages = data.get("totalPages") or data.get("total_pages") or 1

        if isinstance(items, list):
            all_items.extend(items)
        page += 1
        if page >= total_pages:
            break

    sample = all_items[0] if all_items else None
    return _json_response({
        "shop_id": shop_id,
        "url": first_url,
        "raw_first_page_type": type(raw_first_page).__name__,
        "raw_first_page_keys": list(raw_first_page.keys()) if isinstance(raw_first_page, dict) else None,
        "raw_first_page": raw_first_page,
        "total_returns": len(all_items),
        "sample_return_keys": list(sample.keys()) if sample and isinstance(sample, dict) else None,
        "all_returns": all_items,
    })


# ===================================================================
# 20. /debug/return-products
# ===================================================================
@debug_bp.get("/return-products")
@login_required
def debug_return_products():
    """Fetch products of a specific return and return raw JSON for debugging."""
    shop_id = request.args.get("shop_id", "").strip()
    return_id = request.args.get("return_id", "").strip()

    if not shop_id or not return_id:
        return _json_response({"error": "Usage: /debug/return-products?shop_id=10945&return_id=1006608964"}, 400)

    url = f"https://api-seller.uzum.uz/api/seller/shop/{shop_id}/return/{return_id}"
    try:
        data = http_json(url)
    except Exception as e:
        return _json_response({"error": f"API error: {str(e)}"}, 500)

    return _json_response({
        "shop_id": shop_id,
        "return_id": return_id,
        "raw_response_type": type(data).__name__,
        "raw_response_keys": list(data.keys()) if isinstance(data, dict) else None,
        "raw_response": data,
    })


# ===================================================================
# 21. /debug/return-summary
# ===================================================================
@debug_bp.get("/return-summary")
@login_required
def debug_return_summary():
    """Fetch ALL returns, then fetch products for each, and aggregate by skuTitle."""
    shop_id = request.args.get("shop_id", "").strip()
    statuses = request.args.get("statuses", "CREATED,SENT,IN_PROGRESS,MOVED_TO_DELIVERY,ASSEMBLED,COMPLETED,UTILIZED").strip()
    types = request.args.get("types", "FBS,DEFECTED,RETURN").strip()
    page_size = 100

    if not shop_id:
        with SessionLocal() as s:
            shop = s.query(Shop).first()
            if shop:
                shop_id = shop.uzum_id

    if not shop_id:
        return _json_response({"error": "No shop_id provided"}, 400)

    # Step 1: Fetch all returns to get IDs
    all_returns = []
    page = 0
    while True:
        url = (
            f"https://api-seller.uzum.uz/api/seller/shop/{shop_id}/return"
            f"?page={page}&size={page_size}&statuses={statuses}&types={types}"
        )
        try:
            data = http_json(url)
        except Exception as e:
            return _json_response({"error": f"Failed fetching returns page {page}: {str(e)}"}, 500)

        # Unwrap payload if present
        if isinstance(data, dict) and "payload" in data:
            data = data["payload"]

        if isinstance(data, list):
            all_returns.extend(data)
            break
        elif isinstance(data, dict):
            items = data.get("content") or data.get("items") or data.get("data") or []
            if isinstance(items, list):
                all_returns.extend(items)
            total_pages = data.get("totalPages") or data.get("total_pages") or 1
            page += 1
            if page >= total_pages:
                break
        else:
            break

    # Extract return IDs
    return_ids = []
    for ret in all_returns:
        if isinstance(ret, dict):
            rid = ret.get("id") or ret.get("returnId")
            if rid:
                return_ids.append(rid)

    # Step 2: Fetch products for each return IN PARALLEL (10 concurrent)
    from concurrent.futures import ThreadPoolExecutor, as_completed

    summary = {}  # skuTitle -> aggregated data
    errors = []

    def _fetch_return_products(rid):
        url = f"https://api-seller.uzum.uz/api/seller/shop/{shop_id}/return/{rid}"
        try:
            return rid, http_json(url), None
        except Exception as e:
            return rid, None, str(e)

    with ThreadPoolExecutor(max_workers=10) as pool:
        futures = {pool.submit(_fetch_return_products, rid): rid for rid in return_ids}
        for future in as_completed(futures):
            rid, ret_data, err = future.result()
            if err:
                errors.append({"return_id": rid, "error": err})
                continue

            payload = ret_data.get("payload") or {} if isinstance(ret_data, dict) else {}
            products = payload.get("returnItems") or []

            if not isinstance(products, list):
                products = []

            for p in products:
                if not isinstance(p, dict):
                    continue
                sku_title = p.get("skuTitle") or "Unknown"
                qty = p.get("amount") or 0
                price = p.get("purchasePrice") or 0

                if sku_title not in summary:
                    summary[sku_title] = {
                        "skuTitle": sku_title,
                        "totalQuantity": 0,
                        "purchasePrice": price,
                        "returnCount": 0,
                        "totalCost": 0,
                    }
                summary[sku_title]["totalQuantity"] += qty
                summary[sku_title]["returnCount"] += 1
                summary[sku_title]["totalCost"] += qty * price

    sorted_summary = sorted(summary.values(), key=lambda x: x["skuTitle"])
    grand_total_qty = sum(item["totalQuantity"] for item in sorted_summary)
    grand_total_cost = sum(item["totalCost"] for item in sorted_summary)

    return _json_response({
        "shop_id": shop_id,
        "total_returns": len(return_ids),
        "return_ids": return_ids,
        "unique_products": len(sorted_summary),
        "grand_total_quantity": grand_total_qty,
        "grand_total_cost": grand_total_cost,
        "errors": errors,
        "products_summary": sorted_summary,
    })


# ===================================================================
# 22. /debug/return-skus
# ===================================================================
@debug_bp.get("/return-skus")
@login_required
def debug_return_skus():
    """Raw debug for product/return/sku endpoint. Usage: /debug/return-skus?shop_id=5983&product_id=2443455"""
    shop_id = request.args.get("shop_id", "").strip()
    product_id = request.args.get("product_id", "").strip()

    if not shop_id:
        with SessionLocal() as s:
            shop = s.query(Shop).first()
            if shop:
                shop_id = shop.uzum_id

    if not shop_id or not product_id:
        return _json_response({"error": "Usage: /debug/return-skus?shop_id=5983&product_id=2443455"}, 400)

    url = f"https://api-seller.uzum.uz/api/seller/shop/{shop_id}/product/return/sku?productId={product_id}"
    try:
        data = http_json(url)
    except Exception as e:
        return _json_response({"error": f"API error: {str(e)}"}, 500)

    return _json_response({
        "shop_id": shop_id,
        "product_id": product_id,
        "raw_response": data,
    })


# ===================================================================
# 23. /debug/invoice-products
# ===================================================================
@debug_bp.get("/invoice-products")
@login_required
def debug_invoice_products():
    """Fetch products of a specific invoice and return raw JSON for debugging."""
    shop_id = request.args.get("shop_id", "").strip()
    invoice_id = request.args.get("invoice_id", "").strip()

    if not shop_id or not invoice_id:
        return _json_response({"error": "Both shop_id and invoice_id are required. Usage: /debug/invoice-products?shop_id=10945&invoice_id=3300863"}, 400)

    url = f"https://api-seller.uzum.uz/api/seller/shop/{shop_id}/invoice/getInvoiceProducts?invoiceId={invoice_id}"

    try:
        data = http_json(url)
    except Exception as e:
        return _json_response({"error": f"API error: {str(e)}"}, 500)

    # Handle both list and dict responses
    if isinstance(data, list):
        items = data
    elif isinstance(data, dict):
        items = data.get("content") or data.get("items") or data.get("products") or data.get("data") or []
    else:
        items = []

    sample = items[0] if items and isinstance(items, list) else None
    return _json_response({
        "shop_id": shop_id,
        "invoice_id": invoice_id,
        "total_products": len(items) if isinstance(items, list) else None,
        "sample_product_keys": list(sample.keys()) if sample and isinstance(sample, dict) else None,
        "sample_product": sample,
        "all_products": items,
        "raw_response_type": type(data).__name__,
        "raw_response_keys": list(data.keys()) if isinstance(data, dict) else None,
    })


# ===================================================================
# /debug/reports-viewer  — HTML page: shows all report types for a shop
# ===================================================================
@debug_bp.get("/reports-viewer")
@login_required
def debug_reports_viewer():
    """Interactive HTML debug page for Uzum seller documents/reports.
    Usage: /debug/reports-viewer?shop_id=5983
    Each report type is loaded in parallel via JS; clicking fetches CSV data.
    """
    from flask import render_template

    shop_id = request.args.get("shop_id", "").strip()

    # Auto-detect shop_id if not provided
    if not shop_id:
        with SessionLocal() as s:
            shop = s.query(Shop).first()
            if shop:
                shop_id = shop.uzum_id or ""

    # Primary (actively used): SELLS_REPORT and EXPENSES_REPORT
    PRIMARY_JOB_TYPES = [
        "SELLS_REPORT",
        "EXPENSES_REPORT",
    ]
    # Legacy (kept for future use, not auto-loaded)
    LEGACY_JOB_TYPES = [
        "LEFT_OUT_REPORT",
        "MARKED_SALES_REPORT",
        "COMMISSIONER_REPORT",
        "PAID_STORAGE_REPORT",
        "SELLER_RETURN_PAID_STORAGE_REPORT",
        "SELLER_STORAGE_REPORT",
        "LEFT_OUT_REPORT_2024",
    ]

    from datetime import date, timedelta
    today = date.today()
    return render_template(
        "debug_reports_viewer.html",
        shop_id=shop_id,
        primary_job_types=PRIMARY_JOB_TYPES,
        legacy_job_types=LEGACY_JOB_TYPES,
        today=today.strftime("%Y-%m-%d"),
        today_minus_30=(today - timedelta(days=30)).strftime("%Y-%m-%d"),
    )
