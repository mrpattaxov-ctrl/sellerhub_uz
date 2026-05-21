"""Finance-related routes extracted from app.py as a Flask Blueprint."""
from __future__ import annotations

import io
import json
from datetime import date

from flask import Blueprint, render_template, request, send_file
from flask_login import current_user, login_required
from sqlalchemy import func, select

from core.auth_helpers import _current_user_is_admin, _json_response, _user_shop_ids
from core.http_client import http_json
from core.time_helpers import _today_app_tz
from extensions import SessionLocal
from models import FinanceOrder, ProductGroup, Shop, Variant

try:
    import openpyxl
    from openpyxl.styles import Font
except ImportError:
    openpyxl = None


finance_bp = Blueprint("finance_bp", __name__)

_app = None


def init_finance_routes(app_module):
    global _app
    _app = app_module


@finance_bp.get("/finance")
@login_required
def finance_page():
    uid = int(current_user.get_id())
    allowed = _user_shop_ids(uid)
    with SessionLocal() as db:
        shops = (
            db.execute(
                select(Shop).where(Shop.id.in_(allowed)).order_by(Shop.uzum_id)
            ).scalars().all()
            if allowed
            else []
        )
    return render_template(
        "finance_sync.html",
        shops=shops,
        is_admin=_current_user_is_admin(),
    )


@finance_bp.get("/api/finance/debug-fetch")
@login_required
def api_finance_debug_fetch():
    shop_id = request.args.get("shop_id", "5983")
    raw_from = request.args.get("date_from", "").strip()
    raw_to = request.args.get("date_to", "").strip()

    today = _today_app_tz()
    try:
        d_from = date.fromisoformat(raw_from) if raw_from else today.replace(day=1)
    except ValueError:
        d_from = today.replace(day=1)
    try:
        d_to = date.fromisoformat(raw_to) if raw_to else today
    except ValueError:
        d_to = today

    from_ts = _app._app_day_start_ts(d_from)
    to_ts = _app._app_day_end_ts(d_to)
    url = (
        f"https://api-seller.uzum.uz/api/seller/finance/orders"
        f"?shopIds={shop_id}&dateFrom={from_ts}&dateTo={to_ts}"
        f"&group=true&page=0&size=2"
    )

    uz_resp = None
    try:
        uz_resp = http_json(url, headers={"Accept-Language": "uz"})
    except Exception:
        pass

    ru_resp = None
    try:
        ru_resp = http_json(url, headers={"Accept-Language": "ru-RU,ru;q=0.9"})
    except Exception:
        pass

    ru_resp2 = None
    try:
        ru_resp2 = http_json(url, headers={"x-language": "ru", "Accept-Language": "ru"})
    except Exception:
        pass

    def _extract_first_title(resp):
        if not isinstance(resp, dict):
            return None
        order_items = resp.get("orderItems", [])
        if not order_items:
            return None
        first = order_items[0]
        product_title = first.get("productTitle", "")
        sku_items = first.get("items", [])
        first_sku_title = sku_items[0].get("productTitle", "") if sku_items else ""
        return {"productTitle": product_title, "first_sku_productTitle": first_sku_title}

    return _json_response({
        "url": url,
        "date_from": d_from.isoformat(),
        "date_to": d_to.isoformat(),
        "uz_header_Accept-Language_uz": _extract_first_title(uz_resp),
        "ru_header_Accept-Language_ru": _extract_first_title(ru_resp),
        "ru_header_x-language_ru": _extract_first_title(ru_resp2),
    })


@finance_bp.get("/api/finance/data")
@login_required
def api_finance_data():
    shop_id = request.args.get("shop_id", "").strip()
    raw_from = request.args.get("date_from", "").strip()
    raw_to = request.args.get("date_to", "").strip()
    sku_filter = request.args.get("sku", "").strip()
    lang = request.args.get("lang", "ru").strip()

    today = _today_app_tz()
    # Empty/blank date_from means "no lower bound" — show all available
    # history (matches the new UI default). date_to still defaults to today.
    d_from: date | None
    try:
        d_from = date.fromisoformat(raw_from) if raw_from else None
    except ValueError:
        d_from = None
    try:
        d_to = date.fromisoformat(raw_to) if raw_to else today
    except ValueError:
        d_to = today

    uid = int(current_user.get_id())
    allowed = _user_shop_ids(uid)

    with SessionLocal() as db:
        if allowed:
            shops = db.execute(select(Shop).where(Shop.id.in_(allowed))).scalars().all()
            allowed_uzum_ids = [s.uzum_id for s in shops]
        else:
            allowed_uzum_ids = []

        if shop_id and shop_id not in allowed_uzum_ids:
            return _json_response({"error": "Access denied"}, 403)

        target_shops = [shop_id] if shop_id else allowed_uzum_ids
        if not target_shops:
            return _json_response({"items": [], "totals": {}})

        # FinanceOrder.shop_id is varchar (Shop.uzum_id convention). Keep
        # as strings — no int coercion needed.
        target_shop_strs = [str(sid).strip() for sid in target_shops if str(sid).strip()]
        if not target_shop_strs:
            return _json_response({"items": [], "totals": {}})

        # FinanceOrder is per-(shop, day, sku) daily aggregate. Filter by
        # period_from BETWEEN d_from AND d_to (inclusive) — Date column, no
        # tz math needed (the day boundary is already encoded in the row).
        # d_from may be None → "no lower bound" → show all available history.
        where_clauses = [
            FinanceOrder.shop_id.in_(target_shop_strs),
            FinanceOrder.period_from <= d_to,
        ]
        if d_from is not None:
            where_clauses.append(FinanceOrder.period_from >= d_from)
        if sku_filter:
            where_clauses.append(FinanceOrder.sku_title.contains(sku_filter))

        # Server-side GROUP BY — single query. FinanceOrder carries
        # product_title_ru / product_title / image_url / characteristics on
        # the row itself, so we MAX() them into the aggregate (no Variant
        # lookup needed). Variant fallback below for any NULLs.
        stmt = (
            select(
                FinanceOrder.sku_title.label("sku_title"),
                func.max(FinanceOrder.product_id).label("product_id"),
                func.max(FinanceOrder.shop_id).label("shop_id"),
                func.max(FinanceOrder.sku_id).label("sku_id"),
                func.max(FinanceOrder.product_title).label("product_title"),
                func.max(FinanceOrder.product_title_ru).label("product_title_ru"),
                func.max(FinanceOrder.image_url).label("image_url"),
                func.max(FinanceOrder.characteristics).label("characteristics"),
                func.coalesce(func.sum(FinanceOrder.amount), 0).label("amount"),
                func.coalesce(func.sum(FinanceOrder.amount_returns), 0).label("amount_returns"),
                func.coalesce(func.sum(FinanceOrder.sell_price), 0).label("sell_price"),
                func.coalesce(func.sum(FinanceOrder.commission), 0).label("commission"),
                func.coalesce(func.sum(FinanceOrder.seller_profit), 0).label("seller_profit"),
                func.coalesce(func.sum(FinanceOrder.purchase_price), 0).label("purchase_price"),
                func.coalesce(func.sum(FinanceOrder.logistics_fee), 0).label("logistics_fee"),
                func.coalesce(func.sum(FinanceOrder.seller_discount), 0).label("seller_discount"),
                func.count().label("row_count"),
            )
            .where(*where_clauses)
            .group_by(FinanceOrder.sku_title)
            .order_by(func.sum(FinanceOrder.amount).desc())
        )

        rows = db.execute(stmt).all()
        total_row_count = sum(int(r.row_count or 0) for r in rows)

        # Fallback enrichment for any sku_title whose FinanceOrder row had
        # NULL product_title / image_url (defensive — backfill rows from
        # OpenAPI usually carry these). Only looks up SKUs missing both.
        missing_skus = [
            r.sku_title for r in rows
            if r.sku_title and not (r.product_title or r.image_url)
        ]
        variant_meta: dict[str, dict] = {}
        if missing_skus:
            var_rows = db.execute(
                select(
                    Variant.sku,
                    Variant.image_url,
                    Variant.color,
                    Variant.size,
                    ProductGroup.name,
                )
                .join(ProductGroup, Variant.group_id == ProductGroup.id)
                .where(Variant.sku.in_(missing_skus))
            ).all()
            for vr in var_rows:
                bits: list[str] = []
                if vr.color:
                    bits.append(str(vr.color))
                if vr.size:
                    bits.append(str(vr.size))
                variant_meta[vr.sku] = {
                    "product_title": vr.name or "",
                    "image_url": vr.image_url or "",
                    "characteristics": ", ".join(bits),
                }

        items = []
        for r in rows:
            meta = variant_meta.get(r.sku_title, {})
            # Per-language title — finance_orders has both ru + uz titles
            # baked into the row. Fall through to Variant group name if
            # both are NULL.
            if lang == "uz":
                product_title = r.product_title or r.product_title_ru or meta.get("product_title", "")
            else:
                product_title = r.product_title_ru or r.product_title or meta.get("product_title", "")
            image_url = r.image_url or meta.get("image_url", "")
            characteristics = r.characteristics or meta.get("characteristics", "")
            items.append({
                "sku": r.sku_title,
                "product_title": product_title,
                "product_id": r.product_id,
                "image_url": image_url,
                "shop_id": str(r.shop_id) if r.shop_id is not None else "",
                "characteristics": characteristics,
                "amount": int(r.amount or 0),
                "amount_returns": int(r.amount_returns or 0),
                "sell_price": float(r.sell_price or 0),
                "commission": float(r.commission or 0),
                "seller_profit": float(r.seller_profit or 0),
                "purchase_price": float(r.purchase_price or 0),
                "logistics_fee": float(r.logistics_fee or 0),
                "seller_discount": float(r.seller_discount or 0),
            })
        labels = {
            "ru": {
                "amount": "Количество", "returns": "Возвраты", "revenue": "Выручка",
                "commission": "Комиссия", "profit": "Прибыль", "cost": "Себестоимость",
                "logistics": "Логистика", "sku": "Артикул", "product": "Товар",
                "orders": "Заказы", "total": "Итого", "sync": "Синхронизация",
                "full_sync": "Полная синхронизация", "refresh": "Обновить (30 дней)",
                "download": "Скачать Excel", "filter": "Фильтр", "from": "С",
                "to": "По", "search": "Поиск по SKU", "apply": "Применить",
            },
            "uz": {
                "amount": "Miqdori", "returns": "Qaytarishlar", "revenue": "Tushum",
                "commission": "Komissiya", "profit": "Foyda", "cost": "Tannarx",
                "logistics": "Logistika", "sku": "Artikul", "product": "Mahsulot",
                "orders": "Buyurtmalar", "total": "Jami", "sync": "Sinxronlash",
                "full_sync": "To'liq sinxronlash", "refresh": f"Yangilash ({_app.FINANCE_REFRESH_DAYS} kun)",
                "download": "Excel yuklab olish", "filter": "Filtr", "from": "Dan",
                "to": "Gacha", "search": "SKU bo'yicha qidirish", "apply": "Qo'llash",
            },
        }
        labels["ru"]["refresh"] = labels["ru"]["refresh"].replace("30", str(_app.FINANCE_REFRESH_DAYS), 1)

        return _json_response({
            "date_from": d_from.isoformat() if d_from else "",
            "date_to": d_to.isoformat(),
            "total_orders": total_row_count,
            "total_skus": len(items),
            "labels": labels.get(lang, labels["ru"]),
            "totals": {
                "amount": sum(item["amount"] for item in items),
                "amount_returns": sum(item["amount_returns"] for item in items),
                "revenue": sum(item["sell_price"] for item in items),
                "commission": sum(item["commission"] for item in items),
                "profit": sum(item["seller_profit"] for item in items),
                "cost": sum(item["purchase_price"] for item in items),
                "logistics": sum(item["logistics_fee"] for item in items),
            },
            "items": items,
        })


@finance_bp.post("/api/finance/export")
@login_required
def api_finance_export():
    if not openpyxl:
        return "openpyxl not installed", 500

    raw = request.form.get("data")
    if not raw:
        return "No data", 400

    data = json.loads(raw)
    items = data.get("items", [])
    totals = data.get("totals", {})
    labels = data.get("labels", {})

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = labels.get("total", "Finance")

    headers = [
        labels.get("sku", "SKU"),
        labels.get("product", "Product"),
        labels.get("amount", "Amount"),
        labels.get("returns", "Returns"),
        labels.get("revenue", "Revenue"),
        labels.get("commission", "Commission"),
        labels.get("logistics", "Logistics"),
        labels.get("profit", "Profit"),
        labels.get("cost", "Cost"),
    ]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)

    for item in items:
        ws.append([
            item.get("sku", ""),
            item.get("product_title", ""),
            item.get("amount", 0),
            item.get("amount_returns", 0),
            item.get("sell_price", 0),
            item.get("commission", 0),
            item.get("logistics_fee", 0),
            item.get("seller_profit", 0),
            item.get("purchase_price", 0),
        ])

    ws.append([])
    ws.append([
        labels.get("total", "TOTAL"), "",
        totals.get("amount", 0),
        totals.get("amount_returns", 0),
        totals.get("revenue", 0),
        totals.get("commission", 0),
        totals.get("logistics", 0),
        totals.get("profit", 0),
        totals.get("cost", 0),
    ])
    for cell in ws[ws.max_row]:
        cell.font = Font(bold=True)

    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                max_len = max(max_len, len(str(cell.value or "")))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = max(max_len + 2, 12)

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)

    filename = f"finance_{date.today().isoformat()}.xlsx"
    return send_file(
        out,
        download_name=filename,
        as_attachment=True,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
