"""POS and invoice routes extracted from app.py as a Flask Blueprint."""
from __future__ import annotations

import io
import json
from datetime import date, datetime

import debug_routes
from flask import Blueprint, render_template, request, send_file
from flask_login import current_user, login_required
from sqlalchemy import delete, func, select

from core.auth_helpers import _json_response, _user_shop_ids
from core.http_client import http_json
from core.parsers import _safe_qty
from extensions import SessionLocal
from models import PosActionLog, ProductGroup, Shop, Variant, VariantSale

POS_HISTORY_LIMIT = 20

try:
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill
except ImportError:
    openpyxl = None


pos_bp = Blueprint("pos_bp", __name__)

_app = None


def init_pos_routes(app_module):
    global _app
    _app = app_module


@pos_bp.get("/pos")
@login_required
def pos_page():
    return render_template("pos.html")


@pos_bp.get("/api/pos/search")
@login_required
def pos_search():
    query = (request.args.get("q") or "").strip()
    if not query:
        return _json_response({"items": []})

    like = f"%{query}%"
    uid = int(current_user.get_id())
    allowed_shop_ids = _user_shop_ids(uid)
    with SessionLocal() as db:
        stmt = (
            select(Variant, ProductGroup)
            .join(ProductGroup, Variant.group_id == ProductGroup.id)
            .where(
                Variant.sku.ilike(like) |
                Variant.barcode.ilike(like) |
                ProductGroup.name.ilike(like)
            )
        )
        if allowed_shop_ids:
            stmt = stmt.where(ProductGroup.shop_id.in_(allowed_shop_ids))
        else:
            stmt = stmt.where(False)
        stmt = stmt.order_by(
            (ProductGroup.uzum_sort_order == 0).asc(),
            ProductGroup.uzum_sort_order.asc(),
            func.lower(ProductGroup.name).asc(),
            (func.coalesce(Variant.size, "") == "").asc(),
            func.length(func.coalesce(Variant.size, "")).asc(),
            func.lower(func.coalesce(Variant.size, "")).asc(),
            (func.coalesce(Variant.color, "") == "").asc(),
            func.lower(func.coalesce(Variant.color, "")).asc(),
            func.lower(Variant.sku).asc(),
            Variant.id.asc(),
        )
        rows = db.execute(stmt.limit(50)).all()

        items = []
        for variant, group in rows:
            name = group.name
            attrs = []
            if variant.color:
                attrs.append(variant.color)
            if variant.size:
                attrs.append(variant.size)
            if attrs:
                name += f" ({', '.join(attrs)})"

            items.append({
                "id": variant.id,
                "name": name,
                "sku": variant.sku,
                "barcode": variant.barcode,
                "price": variant.price_sum or 0,
                "stock": variant.warehouse_quantity,
                "image_url": variant.image_url or group.image_url,
            })

    return _json_response({"items": items})


def _variant_display_name(variant: Variant) -> str:
    group_name = variant.group.name if variant.group else ""
    attrs = []
    if variant.color:
        attrs.append(variant.color)
    if variant.size:
        attrs.append(variant.size)
    if attrs:
        return f"{group_name} ({', '.join(attrs)})"
    return group_name


def _trim_pos_history(db, user_id: int, keep: int = POS_HISTORY_LIMIT) -> None:
    ids_to_keep = db.execute(
        select(PosActionLog.id)
        .where(PosActionLog.user_id == user_id)
        .order_by(PosActionLog.created_at.desc(), PosActionLog.id.desc())
        .limit(keep)
    ).scalars().all()
    if not ids_to_keep:
        return
    db.execute(
        delete(PosActionLog).where(
            PosActionLog.user_id == user_id,
            ~PosActionLog.id.in_(ids_to_keep),
        )
    )


@pos_bp.post("/api/pos/transaction")
@login_required
def pos_transaction():
    payload = request.get_json(force=True, silent=True) or {}
    mode = payload.get("mode")
    items = payload.get("items") or []

    if not items:
        return _json_response({"error": "No items"}, 400)
    if mode not in ("sale", "stock_in"):
        return _json_response({"error": "Invalid mode"}, 400)

    uid = int(current_user.get_id())

    with SessionLocal() as db:
        snapshot = []
        shop_id = None

        for item in items:
            variant_id = item.get("id")
            qty = int(item.get("qty") or 0)
            if qty <= 0:
                continue

            variant = db.get(Variant, variant_id)
            if not variant:
                continue

            qty_before = int(variant.warehouse_quantity or 0)
            entry = {
                "variant_id": variant.id,
                "sku": variant.sku,
                "name": _variant_display_name(variant),
                "qty": qty,
                "qty_before": qty_before,
            }

            if mode == "sale":
                variant.warehouse_quantity = qty_before - qty
                sale = VariantSale(variant_id=variant.id, date=date.today(), qty_sold=qty)
                db.add(sale)
                db.flush()
                entry["variant_sale_id"] = sale.id
            else:  # stock_in
                variant.warehouse_quantity = qty_before + qty

            entry["qty_after"] = int(variant.warehouse_quantity or 0)

            if shop_id is None and variant.group is not None:
                shop_id = variant.group.shop_id

            snapshot.append(entry)

        if snapshot:
            db.add(PosActionLog(
                user_id=uid,
                shop_id=shop_id,
                action=mode,
                items_json=json.dumps(snapshot, ensure_ascii=False),
            ))
            _trim_pos_history(db, uid)

        db.commit()

    return _json_response({"ok": True})


@pos_bp.get("/api/pos/history")
@login_required
def pos_history():
    uid = int(current_user.get_id())
    with SessionLocal() as db:
        rows = db.execute(
            select(PosActionLog)
            .where(PosActionLog.user_id == uid)
            .order_by(PosActionLog.created_at.desc(), PosActionLog.id.desc())
            .limit(POS_HISTORY_LIMIT)
        ).scalars().all()

        out = []
        for row in rows:
            try:
                entry_items = json.loads(row.items_json) or []
            except (TypeError, ValueError):
                entry_items = []
            total_qty = sum(int(it.get("qty") or 0) for it in entry_items if isinstance(it, dict))
            out.append({
                "id": row.id,
                "action": row.action,
                "created_at": row.created_at.isoformat() + "Z" if row.created_at else None,
                "reverted_at": (row.reverted_at.isoformat() + "Z") if row.reverted_at else None,
                "total_qty": total_qty,
                "items": entry_items,
            })

    return _json_response({"items": out})


@pos_bp.post("/api/pos/undo/<int:action_id>")
@login_required
def pos_undo(action_id: int):
    uid = int(current_user.get_id())

    with SessionLocal() as db:
        log = db.get(PosActionLog, action_id)
        if not log:
            return _json_response({"error": "Not found"}, 404)
        if log.user_id != uid:
            return _json_response({"error": "Forbidden"}, 403)
        if log.reverted_at is not None:
            return _json_response({"error": "Already reverted"}, 409)

        try:
            entries = json.loads(log.items_json) or []
        except (TypeError, ValueError):
            entries = []

        for entry in entries:
            if not isinstance(entry, dict):
                continue
            variant_id = entry.get("variant_id")
            qty = int(entry.get("qty") or 0)
            if not variant_id or qty <= 0:
                continue

            variant = db.get(Variant, variant_id)
            if not variant:
                continue

            if log.action == "sale":
                variant.warehouse_quantity = int(variant.warehouse_quantity or 0) + qty
                sale_id = entry.get("variant_sale_id")
                if sale_id:
                    sale = db.get(VariantSale, sale_id)
                    if sale is not None:
                        db.delete(sale)
            elif log.action == "stock_in":
                variant.warehouse_quantity = int(variant.warehouse_quantity or 0) - qty

        log.reverted_at = datetime.utcnow()
        db.commit()

    return _json_response({"ok": True})


@pos_bp.get("/lost-goods")
@login_required
def lost_goods_page():
    uid = int(current_user.get_id())
    allowed = _user_shop_ids(uid)
    with SessionLocal() as db:
        shops = db.execute(select(Shop).where(Shop.id.in_(allowed))).scalars().all() if allowed else []
    return render_template("lost_goods.html", shops=shops)


@pos_bp.get("/api/lost-goods")
@login_required
def api_lost_goods():
    return debug_routes.debug_lost_goods()


@pos_bp.post("/lost-goods/export")
@login_required
def lost_goods_export():
    if not openpyxl:
        return "openpyxl library not installed", 500

    raw = request.form.get("data")
    if not raw:
        return "No data", 400

    data = json.loads(raw)
    items = data.get("items", [])
    totals = data.get("totals", {})
    shop_id = data.get("shop_id", "")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Lost Goods"

    headers = ["SKU", "Invoiced", "Active (В продаже)", "Defected (Брак)", "Sold", "Lost"]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")

    for item in items:
        ws.append([
            item.get("sku", ""),
            item.get("invoiced", 0),
            item.get("active", 0),
            item.get("defected", 0),
            item.get("sold", 0),
            item.get("lost", 0),
        ])

    ws.append([])
    ws.append([
        "ИТОГО",
        totals.get("invoiced", 0),
        totals.get("active", 0),
        totals.get("defected", 0),
        totals.get("sold", 0),
        totals.get("lost", 0),
    ])
    for cell in ws[ws.max_row]:
        cell.font = Font(bold=True)

    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                max_len = max(max_len, len(str(cell.value or "")))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = max(max_len + 2, 12)

    red_fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=6, max_col=6):
        for cell in row:
            if isinstance(cell.value, (int, float)):
                if cell.value > 0:
                    cell.fill = red_fill
                elif cell.value < 0:
                    cell.fill = yellow_fill

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)

    filename = f"lost_goods_{shop_id}_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    return send_file(
        out,
        download_name=filename,
        as_attachment=True,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@pos_bp.post("/api/pos/fetch-invoice")
@login_required
def pos_fetch_invoice():
    payload = request.get_json(force=True, silent=True) or {}
    shop_id = str(payload.get("shop_id") or "").strip()
    invoice_id = str(payload.get("invoice_id") or "").strip()

    if not shop_id or not invoice_id:
        return _json_response({"error": "Shop ID and Invoice ID are required"}, 400)

    url = f"https://api-seller.uzum.uz/api/seller/shop/{shop_id}/invoice/getInvoiceProducts?invoiceId={invoice_id}"
    try:
        raw_data = http_json(url)
    except Exception as exc:
        return _json_response({"error": f"Failed to fetch invoice from API: {exc}"}, 500)

    items = _app.find_first_array(raw_data, ["items", "products", "rows", "content", "data"]) or []
    if not isinstance(items, list):
        items = []

    all_rows = []
    for item in items:
        nested = item.get("skuForInvoiceDtoList")
        if nested and isinstance(nested, list):
            all_rows.extend(nested)
        else:
            all_rows.append(item)

    found_items = []
    uid = int(current_user.get_id())
    allowed_shop_ids = _user_shop_ids(uid)

    with SessionLocal() as db:
        stmt = select(Variant).join(ProductGroup, Variant.group_id == ProductGroup.id)
        if allowed_shop_ids:
            stmt = stmt.where(ProductGroup.shop_id.in_(allowed_shop_ids))
        else:
            stmt = stmt.where(False)
        variants = db.execute(stmt).scalars().all()
        sku_map = {variant.sku.strip().upper(): variant for variant in variants if variant.sku}
        barcode_map = {variant.barcode.strip().upper(): variant for variant in variants if variant.barcode}

        for item in all_rows:
            invoice_sku = str(_app.pick(item, ["sku", "skuTitle", "shopSku", "offerId"], default="") or "").strip().upper()
            invoice_barcode = str(_app.pick(item, ["barcode", "ean"], default="") or "").strip().upper()

            qty = int(_safe_qty(item.get("quantityToStock")))
            if qty == 0:
                qty = int(_safe_qty(_app.pick(item, ["quantity", "qty", "amount", "count"], default=0)))
            if qty <= 0:
                continue

            matched_variant = None
            if invoice_sku and invoice_sku in sku_map:
                matched_variant = sku_map[invoice_sku]
            elif invoice_barcode and invoice_barcode in barcode_map:
                matched_variant = barcode_map[invoice_barcode]

            if matched_variant:
                group = matched_variant.group
                found_items.append({
                    "id": matched_variant.id,
                    "name": group.name,
                    "sku": matched_variant.sku,
                    "barcode": matched_variant.barcode,
                    "stock": matched_variant.warehouse_quantity,
                    "image_url": matched_variant.image_url or group.image_url,
                    "qty": qty,
                })

    return _json_response({"items": found_items})
