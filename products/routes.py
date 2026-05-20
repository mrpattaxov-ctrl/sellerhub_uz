"""Product/group/sync-related routes extracted from app.py as a Flask Blueprint."""
from __future__ import annotations

import io
import zipfile
from datetime import date, datetime, timedelta, time as dt_time

from flask import Blueprint, redirect, render_template, request, send_file, url_for
from flask_login import current_user, login_required
from sqlalchemy import select, func, delete, update

from extensions import SessionLocal
from models import ProductGroup, Variant, VariantSale, Shop
from core.parsers import _safe_qty
from core.sales_reads import (
    day_bounds_tashkent,
    read_sales_aggregated,
)
from core.time_helpers import _today_app_tz
from core.http_client import http_post_multipart
from core.auth_helpers import (
    _json_response,
    _jwt_expires_in_seconds,
    _get_fresh_api_key,
    _get_admin_token,
    _current_user_is_admin,
    admin_required,
    _user_shop_ids,
)

try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, Border, Side
except ImportError:
    openpyxl = None

products_bp = Blueprint("products_bp", __name__)

# ── Late-binding to avoid circular imports ──────────────────────────────
_app = None


def init_products_routes(app_module):
    global _app
    _app = app_module


# ── Routes ──────────────────────────────────────────────────────────────

@products_bp.get("/")
@login_required
def home():
    return redirect(url_for("products_bp.groups_page"))

@products_bp.get("/print/labels")
@login_required
def print_labels():
    ids_str = request.args.get("ids") or ""
    if not ids_str:
        return "No items selected", 400

    try:
        ids = [int(x) for x in ids_str.split(",") if x.strip().isdigit()]
    except ValueError:
        return "Invalid IDs", 400

    LABEL_SIZES = {
        "30x20":  {"w": 30, "h": 20, "tcol": 7,  "qr": 14, "sku_fs": 5.4, "num_fs": 6, "num_last4_fs": 8},
        "43x25":  {"w": 43, "h": 25, "tcol": 8,  "qr": 20, "sku_fs": 6,   "num_fs": 7, "num_last4_fs": 9},
        "40x30":  {"w": 40, "h": 30, "tcol": 9,  "qr": 20, "sku_fs": 6.5, "num_fs": 7, "num_last4_fs": 9},
        "60x60":  {"w": 60, "h": 60, "tcol": 12, "qr": 34, "sku_fs": 8,   "num_fs": 8, "num_last4_fs": 11},
        "70x37":  {"w": 70, "h": 37, "tcol": 12, "qr": 40, "sku_fs": 8,   "num_fs": 8, "num_last4_fs": 11},
    }
    size_key = request.args.get("size", "30x20")
    if size_key not in LABEL_SIZES:
        size_key = "30x20"
    lbl = LABEL_SIZES[size_key]

    with SessionLocal() as db:
        # Fetch unique variants first
        unique_ids = list(set(ids))
        if not unique_ids:
             return "No items", 400
        objs = db.execute(select(Variant).where(Variant.id.in_(unique_ids))).scalars().all()
        obj_map = {o.id: o for o in objs}

        # Rebuild list with duplicates based on input 'ids' order to support quantity
        variants = []
        for i in ids:
            if i in obj_map:
                variants.append(obj_map[i])

    return render_template("print_labels.html", variants=variants, lbl=lbl)

@products_bp.get("/print/queue")
@login_required
def print_queue_page():
    return render_template("print_queue.html")

@products_bp.get("/groups")
@login_required
def groups_page():
    q = (request.args.get("q") or "").strip()
    shop_filter = (request.args.get("shop_id") or "").strip()
    status_filter = (request.args.get("status") or "active").strip().lower()
    display_status = "archived" if status_filter in ("archived", "archive") else "active"
    page = max(1, int(request.args.get("page") or 1))
    per_page = 50

    # Restrict to the user's assigned shops
    uid = int(current_user.get_id())
    allowed_shop_ids = _user_shop_ids(uid)

    with SessionLocal() as db:
        stmt = select(ProductGroup)

        # Always filter to allowed shops; also include groups with no shop assigned (NULL)
        if allowed_shop_ids:
            stmt = stmt.where(
                ProductGroup.shop_id.in_(allowed_shop_ids) | (ProductGroup.shop_id == None)
            )
        else:
            stmt = stmt.where(ProductGroup.shop_id == None)

        if shop_filter and shop_filter.isdigit() and int(shop_filter) in allowed_shop_ids:
            stmt = stmt.where(ProductGroup.shop_id == int(shop_filter))

        if display_status == "archived":
            stmt = stmt.where(ProductGroup.is_archived == True)
        else:
            stmt = stmt.where((ProductGroup.is_archived == False) | (ProductGroup.is_archived == None))

        if q:
            like = f"%{q}%"
            # Use a subquery to avoid SELECT DISTINCT + ORDER BY expression conflict in PostgreSQL
            subq = stmt.outerjoin(ProductGroup.variants).where(
                ProductGroup.name.ilike(like) |
                Variant.sku.ilike(like) |
                Variant.barcode.ilike(like)
            ).with_only_columns(ProductGroup.id).distinct().subquery()
            stmt = select(ProductGroup).where(ProductGroup.id.in_(select(subq)))

        # Sort by sku-list position (0 = not in sku-list, goes last), then by id
        stmt = stmt.order_by(
            (ProductGroup.uzum_sort_order == 0).asc(),
            ProductGroup.uzum_sort_order.asc(),
            ProductGroup.id.asc(),
        )

        total_count = db.execute(select(func.count()).select_from(stmt.subquery())).scalar() or 0
        total_pages = max(1, (total_count + per_page - 1) // per_page)
        page = min(page, total_pages)

        groups = db.execute(stmt.offset((page - 1) * per_page).limit(per_page)).scalars().all()

        # aggregate counts only for current page
        group_ids = [g.id for g in groups]
        vstmt = (
            select(
                Variant.group_id,
                func.count(Variant.id),
                func.coalesce(func.sum(Variant.uzum_quantity), 0),
                func.coalesce(func.sum(Variant.warehouse_quantity), 0),
                func.min(Variant.sku),
            )
            .where(Variant.group_id.in_(group_ids) if group_ids else False)
            .group_by(Variant.group_id)
        )
        agg = {gid: {"variants": c, "uzum_qty": int(u), "wh_qty": int(w), "sku": "-".join(str(s).split("-")[:2]) if s else ""} for (gid, c, u, w, s) in db.execute(vstmt).all()}

        # Fetch shops for the picker
        shops = db.execute(
            select(Shop).where(Shop.id.in_(allowed_shop_ids)) if allowed_shop_ids else select(Shop)
        ).scalars().all()

    return render_template("groups.html", groups=groups, agg=agg, q=q,
                           current_status=display_status, shops=shops, current_shop_id=shop_filter,
                           page=page, total_pages=total_pages, total_count=total_count, per_page=per_page)

@products_bp.get("/fetch")
@login_required
def fetch_page():
    return render_template("fetch.html")

@products_bp.get("/groups/<int:group_id>")
@login_required
def group_detail(group_id: int):
    uid = int(current_user.get_id())
    allowed_shop_ids = _user_shop_ids(uid)
    with SessionLocal() as db:
        group = db.get(ProductGroup, group_id)
        if not group:
            return render_template("not_found.html", message="Product not found"), 404
        if not _current_user_is_admin() and group.shop_id not in allowed_shop_ids:
            return render_template("not_found.html", message="Product not found"), 404

        variants = db.execute(
            select(Variant).where(Variant.group_id == group_id).order_by(func.lower(Variant.sku))
        ).scalars().all()

        # 30d sales from sales_lines (Tashkent window [today-30, today+1)).
        # Old path filtered `period_from >= d_from AND period_to <= today` — the new
        # equivalent is `created_at >= today-30d 00:00 AND created_at < today+1d 00:00`.
        shop = db.get(Shop, group.shop_id)
        sales_30d_map: dict[int, int] = {}
        if shop:
            today = _today_app_tz()
            start_ts, _ = day_bounds_tashkent(today - timedelta(days=30))
            _, end_ts = day_bounds_tashkent(today)
            agg_rows = read_sales_aggregated(
                shop.uzum_id,
                start_ts,
                end_ts,
                group_by="sku",
                session=db,
            )
            # Build lookups (assign, don't accumulate — avoids double-counting).
            # `sales_lines.sku_id` carries the seller SKU code (e.g. "LUXUZ-RING-СИНИЙ-17"),
            # which matches `Variant.sku` — NOT the numeric `Variant.uzum_sku_id`.
            by_title: dict[str, int] = {}
            by_sku_id: dict[str, int] = {}
            for row in agg_rows:
                title = (row.get("sku_title") or "").strip()
                qty = int(row.get("qty_sum") or 0)
                if title:
                    by_title[title] = qty
                    by_title[title.upper()] = qty
                sid = row.get("sku_id")
                if sid:
                    sid_s = str(sid)
                    by_sku_id[sid_s] = qty
                    by_sku_id[sid_s.upper()] = qty
            for v in variants:
                vsku = v.sku or ""
                matched = by_sku_id.get(vsku) or by_sku_id.get(vsku.upper()) or 0
                if matched == 0 and v.uzum_sku_id:
                    matched = by_sku_id.get(v.uzum_sku_id, 0)
                if matched == 0:
                    matched = by_title.get(vsku) or by_title.get(vsku.upper()) or 0
                sales_30d_map[v.id] = matched

    return render_template("group_detail.html", group=group, variants=variants,
                           sales_30d_map=sales_30d_map)


@products_bp.get("/economics")
@login_required
def economics_page():
    return render_template("economics.html")


@products_bp.get("/api/economics/data")
@login_required
def economics_data_api():
    """Returns economics data for the selected date range from local finance_orders DB."""
    today = _today_app_tz()
    raw_from = request.args.get("date_from", "").strip()
    raw_to   = request.args.get("date_to",   "").strip()
    try:
        date_from = date.fromisoformat(raw_from) if raw_from else today.replace(day=1)
    except ValueError:
        date_from = today.replace(day=1)
    try:
        date_to = date.fromisoformat(raw_to) if raw_to else today
    except ValueError:
        date_to = today

    uid = int(current_user.get_id())
    allowed_shop_ids = _user_shop_ids(uid)

    with SessionLocal() as db:
        # Get uzum_ids for allowed shops + mapping from internal shop_id to uzum_id
        shops = db.execute(select(Shop).where(Shop.id.in_(allowed_shop_ids))).scalars().all() if allowed_shop_ids else []
        shop_uzum_ids = [s.uzum_id for s in shops]
        shop_id_to_uzum = {s.id: s.uzum_id for s in shops}

        # Aggregate sales_lines per (shop_uzum_id, sku) for the Tashkent window
        # [date_from 00:00, date_to+1 00:00). Old path filtered
        # `period_from >= date_from AND period_to <= date_to` — the new path
        # is the equivalent right-open `created_at` window.
        per_shop_sales: dict[tuple[str, str], dict] = {}
        if shop_uzum_ids:
            start_ts, _ = day_bounds_tashkent(date_from)
            _, end_ts = day_bounds_tashkent(date_to)
            agg_rows = read_sales_aggregated(
                shop_uzum_ids,
                start_ts,
                end_ts,
                group_by="sku",
                session=db,
            )
            # `sales_lines.shop_id` is int but the rest of this module keys
            # lookups by the string uzum_id — cast back to str.
            for row in agg_rows:
                sid = str(row.get("shop_id"))
                title = (row.get("sku_title") or "").strip()
                qty = int(row.get("qty_sum") or 0)
                sell = int(row.get("revenue_sum") or 0)
                comm = int(row.get("commission_sum") or 0)
                logi = int(row.get("logistics_sum") or 0)
                entry = {"qty": qty, "sell_price": sell,
                         "commission": comm, "logistics": logi}
                if title:
                    per_shop_sales[(sid, title)] = entry
                    per_shop_sales[(sid, title.upper())] = entry
                if row.get("sku_id"):
                    per_shop_sales[(sid, str(row["sku_id"]))] = entry

        stmt = select(ProductGroup).where(ProductGroup.is_archived == False)
        if allowed_shop_ids:
            stmt = stmt.where(ProductGroup.shop_id.in_(allowed_shop_ids))
        else:
            stmt = stmt.where(False)
        groups = db.execute(stmt).scalars().all()

        items = []
        t_stock_cost = t_stock_qty = t_sales_rev = 0
        t_sales_qty  = t_sales_profit = t_commission = t_logistics = 0

        for g in groups:
            g_stock_qty = g_stock_cost = g_sales_qty = 0
            g_sales_rev = g_sales_cost = g_commission = g_logistics = 0
            g_uzum_id = shop_id_to_uzum.get(g.shop_id, "")

            for v in g.variants:
                cost      = v.purchase_price or 0
                stock_qty = (v.uzum_quantity or 0) + (v.warehouse_quantity or 0)
                g_stock_qty  += stock_qty
                g_stock_cost += stock_qty * cost

                fin = None
                for key in [v.sku, (v.sku or "").upper(), v.barcode,
                             (v.barcode or "").upper(), v.uzum_sku_id]:
                    if key and (g_uzum_id, key) in per_shop_sales:
                        fin = per_shop_sales[(g_uzum_id, key)]
                        break

                sq              = fin["qty"]              if fin else 0
                # finance_orders stores period totals (not per-unit), so use directly
                total_sell      = fin.get("sell_price", 0) if fin else 0
                total_comm      = fin.get("commission",  0) if fin else 0
                total_logi      = fin.get("logistics",   0) if fin else 0

                g_sales_qty  += sq
                g_sales_rev  += total_sell
                g_sales_cost += sq * cost
                g_commission += total_comm
                g_logistics  += total_logi

            g_sales_profit = g_sales_rev - g_sales_cost - g_commission - g_logistics
            roi = round(g_sales_profit / g_sales_cost * 100, 1) if g_sales_cost > 0 else 0

            items.append({
                "id": g.id, "name": g.name, "image_url": g.image_url or "",
                "stock_qty": g_stock_qty, "stock_cost": g_stock_cost,
                "sales_qty": g_sales_qty, "sales_revenue": g_sales_rev,
                "sales_cost": g_sales_cost, "sales_commission": g_commission,
                "sales_logistics": g_logistics, "sales_profit": g_sales_profit,
                "roi": roi,
            })
            t_stock_cost   += g_stock_cost;  t_stock_qty    += g_stock_qty
            t_sales_rev    += g_sales_rev;   t_sales_qty    += g_sales_qty
            t_sales_profit += g_sales_profit; t_commission   += g_commission
            t_logistics    += g_logistics

        items.sort(key=lambda x: x["sales_profit"], reverse=True)

        return _json_response({
            "items": items,
            "totals": {
                "stock_cost": t_stock_cost, "stock_qty": t_stock_qty,
                "sales_revenue": t_sales_rev, "sales_qty": t_sales_qty,
                "sales_profit": t_sales_profit, "sales_commission": t_commission,
                "sales_logistics": t_logistics,
            },
        })

@products_bp.get("/calculator")
@login_required
def calculator_page():
    return render_template("calculator.html")


# ----------------------------
# Product pages (requested)
# ----------------------------
@products_bp.get("/products")
@login_required
def products_redirect():
    # alias to match "main products page"
    return redirect(url_for("products_bp.groups_page"))


# ----------------------------
# Uzum sync API (new)
# ----------------------------
@products_bp.post("/api/uzum/sync")
@login_required
def uzum_sync():
    try:
        return _uzum_sync_inner()
    except Exception as e:
        import traceback; traceback.print_exc()
        return _json_response({"error": str(e)}, 500)


@products_bp.post("/api/uzum/sync-all")
@login_required
@admin_required
def uzum_sync_all():
    """Discover all seller shops automatically and sync all of them."""
    try:
        result = _app._sync_all_seller_shops()
        return _json_response({"ok": True, **result})
    except Exception as e:
        import traceback; traceback.print_exc()
        return _json_response({"error": str(e)}, 500)


def _uzum_sync_inner():
    payload = request.get_json(force=True, silent=True) or {}
    shop_id = str(payload.get("shop_id") or "").strip()

    if not shop_id:
        return _json_response({"error": "shop_id missing"}, 400)

    if not _current_user_is_admin():
        uid = int(current_user.get_id())
        with SessionLocal() as _db:
            existing = _db.execute(select(Shop).where(Shop.uzum_id == shop_id)).scalar_one_or_none()
        if not existing or existing.owner_id != uid:
            return _json_response({"error": "Access denied to this shop"}, 403)

    if not _get_admin_token():
        return _json_response({"error": "Uzum \u0442\u043e\u043a\u0435\u043d \u043d\u0435 \u043d\u0430\u0441\u0442\u0440\u043e\u0435\u043d."}, 401)

    size = int(payload.get("size") or 100)
    sync_all = bool(payload.get("sync_all", True))
    max_pages = int(payload.get("max_pages") or 500)

    result = _app._sync_products_for_shop(shop_id,
                                     size=size, sync_all=sync_all, max_pages=max_pages)
    return _json_response({"ok": True, "shop_id": shop_id, **result})


@products_bp.post("/api/uzum/sync-finance")
@login_required
def uzum_sync_finance():
    try:
        payload = request.get_json(force=True, silent=True) or {}
        shop_id = str(payload.get("shop_id") or "").strip()
        api_key = _get_admin_token()

        if not shop_id:
            return _json_response({"error": "shop_id missing"}, 400)

        with SessionLocal() as db:
            # Ensure shop exists
            shop_obj = db.execute(select(Shop).where(Shop.uzum_id == shop_id)).scalar_one_or_none()
            if not shop_obj:
                 return _json_response({"error": "\u041c\u0430\u0433\u0430\u0437\u0438\u043d \u043d\u0435 \u043d\u0430\u0439\u0434\u0435\u043d. \u0421\u043d\u0430\u0447\u0430\u043b\u0430 \u0432\u044b\u043f\u043e\u043b\u043d\u0438\u0442\u0435 \u0441\u0438\u043d\u0445\u0440\u043e\u043d\u0438\u0437\u0430\u0446\u0438\u044e \u0442\u043e\u0432\u0430\u0440\u043e\u0432."}, 404)

            # Fetch 30-day sales via the SELLS_REPORT pipeline (one Uzum
            # /documents/v2 call replaces the legacy paginated /finance/orders).
            today = _today_app_tz()
            window_from = datetime.combine(today - timedelta(days=30), dt_time(0, 0, 0))
            window_to   = datetime.combine(today + timedelta(days=1),  dt_time(0, 0, 0))

            try:
                rows = _app._fetch_sells_report_rows_for_shops(
                    [int(shop_id)], window_from, window_to,
                ) or []
            except Exception:
                return _json_response({"error": "\u041d\u0435 \u0443\u0434\u0430\u043b\u043e\u0441\u044c \u043f\u043e\u043b\u0443\u0447\u0438\u0442\u044c \u0434\u0430\u043d\u043d\u044b\u0435 \u043e \u043f\u0440\u043e\u0434\u0430\u0436\u0430\u0445 (\u043e\u0448\u0438\u0431\u043a\u0430 API). \u041f\u0440\u043e\u0432\u0435\u0440\u044c\u0442\u0435 ID \u043c\u0430\u0433\u0430\u0437\u0438\u043d\u0430."}, 500)

            # Seed sales_lines for this 30-day window so the Finance page is
            # populated immediately. Failure here must not abort the variant
            # update \u2014 the next HH:00 hourly bulk will catch up.
            try:
                _app._ingest_sales_lines_window(rows, window_from, window_to)
            except Exception as e:
                print(f"[ProductsSync] sales_lines seed failed for shop={shop_id}: {e!r}")

            # Aggregate per SKU, dropping cancelled rows. Build the same
            # {sku: {qty, price, sell_price, commission, logistics}} shape
            # the variant loop below expects (per-unit values, int).
            agg: dict[str, dict[str, float]] = {}
            for r in rows:
                if (r.get("status") or "").strip() == "\u041e\u0442\u043c\u0435\u043d\u0435\u043d":
                    continue
                sku = str(r.get("sku_id") or "").strip()
                if not sku:
                    continue
                try:
                    q = int(r.get("qty") or 0)
                except (TypeError, ValueError):
                    q = 0
                if q <= 0:
                    continue
                a = agg.setdefault(sku, {
                    "qty": 0,
                    "purchase_price_total": 0.0,
                    "revenue_total": 0.0,
                    "commission_total": 0.0,
                    "logistics_total": 0.0,
                })
                a["qty"] += q
                try:
                    a["purchase_price_total"] += float(r.get("purchase_price") or 0)
                except (TypeError, ValueError):
                    pass
                try:
                    a["revenue_total"] += float(r.get("revenue") or 0)
                except (TypeError, ValueError):
                    pass
                try:
                    a["commission_total"] += float(r.get("commission") or 0)
                except (TypeError, ValueError):
                    pass
                try:
                    a["logistics_total"] += float(r.get("logistics_fee") or 0)
                except (TypeError, ValueError):
                    pass

            sales_map: dict[str, dict] = {}
            for sku, a in agg.items():
                qty = int(a["qty"])
                if qty <= 0:
                    continue
                sales_map[sku] = {
                    "qty": qty,
                    "price":      int(a["purchase_price_total"] / qty) if a["purchase_price_total"] > 0 else 0,
                    "sell_price": int(a["revenue_total"]         / qty) if a["revenue_total"]         > 0 else 0,
                    "commission": int(a["commission_total"]      / qty) if a["commission_total"]      > 0 else 0,
                    "logistics":  int(a["logistics_total"]       / qty) if a["logistics_total"]       > 0 else 0,
                }

            # Update ALL variants in DB for this shop
            variants = db.execute(select(Variant).join(ProductGroup).where(ProductGroup.shop_id == shop_obj.id)).scalars().all()
            updated_count = 0
            _today = date.today()

            # Bulk-delete today's VariantSale records for this shop's variants upfront
            variant_ids = [v.id for v in variants]
            if variant_ids:
                db.execute(delete(VariantSale).where(
                    VariantSale.variant_id.in_(variant_ids),
                    VariantSale.date == _today
                ))

            for v in variants:
                sku_key = (v.sku or "").strip()
                data = sales_map.get(sku_key) or sales_map.get(sku_key.upper())
                if data is None and v.barcode:
                    bc_key = v.barcode.strip()
                    data = sales_map.get(bc_key) or sales_map.get(bc_key.upper())

                qty_val = data["qty"] if data else 0
                v.sales_30d_finance = qty_val
                v.avg_daily_sales = qty_val / 30.0
                if data and data.get("price", 0) > 0:
                    v.purchase_price = data["price"]
                if data and data.get("sell_price", 0) > 0:
                    v.sell_price_uzum = int(data["sell_price"])
                if data and data.get("commission", 0) > 0:
                    v.commission_per_unit = int(data["commission"])
                if data and data.get("logistics", 0) > 0:
                    v.logistics_per_unit = int(data["logistics"])

                # Insert today's VariantSale so economics date-range queries work
                if qty_val > 0:
                    db.add(VariantSale(variant_id=v.id, date=_today, qty_sold=qty_val))
                updated_count += 1

            db.commit()

        return _json_response({"ok": True, "updated": updated_count, "sales_records": len(sales_map)})
    except Exception as e:
        return _json_response({"error": f"\u041e\u0448\u0438\u0431\u043a\u0430 \u0441\u0435\u0440\u0432\u0435\u0440\u0430: {str(e)}"}, 500)

@products_bp.get("/api/groups/<int:group_id>/sales-range")
@login_required
def group_sales_range(group_id: int):
    """Return per-variant sales for a custom date range from local finance_orders DB."""
    days_param = request.args.get("days")
    date_from_str = request.args.get("date_from")
    date_to_str = request.args.get("date_to")

    today = date.today()
    if date_from_str and date_to_str:
        try:
            d_from = date.fromisoformat(date_from_str)
            d_to = date.fromisoformat(date_to_str)
        except ValueError:
            return _json_response({"error": "Invalid date format. Use YYYY-MM-DD."}, 400)
        days_label = (d_to - d_from).days + 1
    else:
        days_label = int(days_param) if days_param else 30
        d_to = today
        d_from = today - timedelta(days=days_label)

    with SessionLocal() as db:
        group = db.get(ProductGroup, group_id)
        if not group:
            return _json_response({"error": "Group not found"}, 404)

        uid = int(current_user.get_id())
        allowed = _user_shop_ids(uid)
        if group.shop_id not in allowed:
            return _json_response({"error": "Access denied"}, 403)

        shop = db.get(Shop, group.shop_id)
        if not shop:
            return _json_response({"error": "Shop not found"}, 404)

        variants = db.execute(
            select(Variant).where(Variant.group_id == group_id)
        ).scalars().all()

        # Aggregate sales_lines by SKU for this shop + date range. Old path
        # `period_from >= d_from AND period_to <= d_to` → right-open
        # `created_at >= d_from 00:00 AND created_at < d_to+1d 00:00`
        # (Tashkent).
        start_ts, _ = day_bounds_tashkent(d_from)
        _, end_ts = day_bounds_tashkent(d_to)
        agg_rows = read_sales_aggregated(
            shop.uzum_id,
            start_ts,
            end_ts,
            group_by="sku",
            session=db,
        )
        # Build lookups. `sales_lines.sku_id` is the seller SKU code that
        # matches `Variant.sku` directly — title / uzum_sku_id are fallbacks.
        sales_by_title: dict[str, int] = {}
        sales_by_sku_id: dict[str, int] = {}
        for row in agg_rows:
            title = (row.get("sku_title") or "").strip()
            qty = int(row.get("qty_sum") or 0)
            if title:
                sales_by_title[title] = qty
                sales_by_title[title.upper()] = qty
            sid = row.get("sku_id")
            if sid:
                sid_s = str(sid)
                sales_by_sku_id[sid_s] = qty
                sales_by_sku_id[sid_s.upper()] = qty

        result = []
        for v in variants:
            vsku = v.sku or ""
            qty = sales_by_sku_id.get(vsku) or sales_by_sku_id.get(vsku.upper()) or 0
            if qty == 0 and v.uzum_sku_id and v.uzum_sku_id in sales_by_sku_id:
                qty = sales_by_sku_id[v.uzum_sku_id]
            if qty == 0:
                for key in [vsku, vsku.upper(), v.barcode,
                            (v.barcode or "").upper()]:
                    if key and key in sales_by_title:
                        qty = sales_by_title[key]
                        break
            result.append({"variant_id": v.id, "qty": qty})

    return _json_response({"sales": result, "days": days_label,
                           "date_from": date_from_str, "date_to": date_to_str})


@products_bp.get("/api/groups/<int:group_id>/variants")
@login_required
def get_group_variants_api(group_id: int):
    with SessionLocal() as db:
        variants = db.execute(
            select(Variant).where(Variant.group_id == group_id).order_by(func.lower(Variant.sku))
        ).scalars().all()

        items = []
        for v in variants:
            s30 = v.sales_30d_finance or 0
            stock = (v.uzum_quantity or 0) + (v.warehouse_quantity or 0)
            need = (s30 * 2) - stock
            items.append({
                "id": v.id,
                "sku": v.sku,
                "image_url": v.image_url,
                "sales_30d": s30,
                "need_60d": need
            })

        return _json_response({
            "variants": items
        })


@products_bp.post("/api/variants/<int:variant_id>/sales")
@login_required
def add_variant_sale(variant_id: int):
    payload = request.get_json(force=True, silent=True) or {}
    try:
        qty = int(payload.get("qty") or 0)
    except ValueError:
        qty = 0
    if qty <= 0:
        return _json_response({"error": "qty must be > 0"}, 400)

    try:
        sale_date = payload.get("date")
        d = date.fromisoformat(sale_date) if sale_date else date.today()
    except Exception:
        d = date.today()

    with SessionLocal() as db:
        v = db.get(Variant, variant_id)
        if not v:
            return _json_response({"error": "Variant not found"}, 404)
        s = VariantSale(variant_id=variant_id, date=d, qty_sold=qty)
        db.add(s)
        db.commit()
        return _json_response({"ok": True})


# ----------------------------
# Invoice / Restock Logic
# ----------------------------
@products_bp.get("/invoice/restock")
@login_required
def invoice_restock_page():
    shop_filter = (request.args.get("shop_id") or "").strip()
    uid = int(current_user.get_id())
    allowed_shop_ids = _user_shop_ids(uid)

    with SessionLocal() as db:
        if _current_user_is_admin():
            shops = db.execute(select(Shop)).scalars().all()
        else:
            shops = db.execute(select(Shop).where(Shop.id.in_(allowed_shop_ids))).scalars().all()

        stmt = select(Variant, ProductGroup).join(ProductGroup, Variant.group_id == ProductGroup.id)
        if allowed_shop_ids:
            stmt = stmt.where(ProductGroup.shop_id.in_(allowed_shop_ids))
        else:
            stmt = stmt.where(False)
        if shop_filter and shop_filter.isdigit() and int(shop_filter) in allowed_shop_ids:
            stmt = stmt.where(ProductGroup.shop_id == int(shop_filter))

        rows = db.execute(stmt).all()

        items = []
        for v, g in rows:
            s30 = v.sales_30d_finance or 0
            u_qty = v.uzum_quantity or 0
            wh_qty = v.warehouse_quantity or 0

            # Logic: needed = s30. If u_qty < needed, restock = needed - u_qty
            if u_qty < s30:
                needed = s30 - u_qty
                if wh_qty > 0:
                    restock = min(needed, wh_qty)
                    price = v.purchase_price or 0
                    items.append({
                        "id": v.id,
                        "name": g.name,
                        "sku": v.sku,
                        "barcode": v.barcode,
                        "sales_30d": s30,
                        "uzum_qty": u_qty,
                        "wh_qty": wh_qty,
                        "restock_qty": restock,
                        "price": price,
                        "total_price": restock * price,
                        "image_url": v.image_url or g.image_url
                    })

        # Sort by SKU to keep variants together
        items.sort(key=lambda x: str(x.get("sku") or "").strip().lower())

        # Chunk into max 35 items per file/invoice
        chunk_size = 35
        chunks = [items[i:i + chunk_size] for i in range(0, len(items), chunk_size)]
        if not chunks:
            chunks = [[]]

    return render_template("invoice_restock.html", chunks=chunks, shops=shops, current_shop=shop_filter)

@products_bp.route("/invoice/restock/download", methods=["GET", "POST"])
@login_required
def invoice_restock_download():
    if not openpyxl:
        return _json_response({"error": "openpyxl library not installed. Please run: pip install openpyxl"}, 500)

    try:
        data_rows = []

        if request.method == "POST":
            # Use data provided by the client (edited quantities)
            payload = request.get_json(force=True, silent=True) or {}
            items = payload.get("items") or []

            # Sort items by SKU to ensure they are grouped nicely in the Excel file
            items.sort(key=lambda x: str(x.get("sku") or "").strip().lower())

            for item in items:
                bc = str(item.get("barcode") or "").strip()
                try:
                    price = float(item.get("price") or 0)
                    qty = int(item.get("qty") or 0)
                except (ValueError, TypeError):
                    continue
                if qty > 0:
                    data_rows.append([bc, price, qty])
        else:
            # GET request: Auto-calculate based on DB (legacy behavior)
            shop_filter = (request.args.get("shop_id") or "").strip()
            uid = int(current_user.get_id())
            allowed_shop_ids = _user_shop_ids(uid)
            with SessionLocal() as db:
                stmt = select(Variant, ProductGroup).join(ProductGroup, Variant.group_id == ProductGroup.id)
                if allowed_shop_ids:
                    stmt = stmt.where(ProductGroup.shop_id.in_(allowed_shop_ids))
                else:
                    stmt = stmt.where(False)
                if shop_filter and shop_filter.isdigit() and int(shop_filter) in allowed_shop_ids:
                    stmt = stmt.where(ProductGroup.shop_id == int(shop_filter))
                stmt = stmt.order_by(Variant.sku)
                rows = db.execute(stmt).all()

                for v, g in rows:
                    s30 = v.sales_30d_finance or 0
                    u_qty = v.uzum_quantity or 0
                    wh_qty = v.warehouse_quantity or 0

                    if u_qty < s30:
                        needed = s30 - u_qty
                        if wh_qty > 0:
                            restock = min(needed, wh_qty)
                            price = v.purchase_price or 0
                            data_rows.append([v.barcode or "", price, restock])

        # Chunk into max 35 items per file
        chunk_size = 35
        chunks = [data_rows[i:i + chunk_size] for i in range(0, len(data_rows), chunk_size)]
        if not chunks:
            chunks = [[]]

        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')

        def create_wb(rows_subset):
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "\u0422\u043e\u0432\u0430\u0440\u044b \u043d\u0430 \u043e\u0442\u043f\u0440\u0430\u0432\u043a\u0443"
            ws.append(["\u0428\u0442\u0440\u0438\u0445\u043a\u043e\u0434 \u0442\u043e\u0432\u0430\u0440\u0430*", "\u0421\u0435\u0431\u0435\u0441\u0442\u043e\u0438\u043c\u043e\u0441\u0442\u044c (\u0441\u0443\u043c)*", "\u041a\u043e\u043b\u0438\u0447\u0435\u0441\u0442\u0432\u043e (\u0448\u0442)*"])
            for cell in ws[1]: cell.font = Font(bold=True)
            for r in rows_subset:
                ws.append(r)
            out = io.BytesIO()
            wb.save(out)
            out.seek(0)
            return out

        if len(chunks) == 1:
            out = create_wb(chunks[0])
            filename = f"invoice_restock_{timestamp}.xlsx"
            mimetype = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            try:
                return send_file(out, download_name=filename, as_attachment=True, mimetype=mimetype)
            except TypeError:
                # Fallback for older Flask versions
                return send_file(out, attachment_filename=filename, as_attachment=True, mimetype=mimetype)
        else:
            zip_buffer = io.BytesIO()
            with zipfile.ZipFile(zip_buffer, "w", zipfile.ZIP_DEFLATED) as zf:
                for i, chunk in enumerate(chunks):
                    xlsx_io = create_wb(chunk)
                    zf.writestr(f"invoice_restock_{timestamp}_part{i+1}.xlsx", xlsx_io.getvalue())

            zip_buffer.seek(0)
            filename = f"invoice_restock_{timestamp}_multi.zip"
            mimetype = "application/zip"
            try:
                return send_file(zip_buffer, download_name=filename, as_attachment=True, mimetype=mimetype)
            except TypeError:
                return send_file(zip_buffer, attachment_filename=filename, as_attachment=True, mimetype=mimetype)

    except Exception as e:
        return _json_response({"error": f"Server Error: {str(e)}"}, 500)


@products_bp.route("/invoice/restock/upload-uzum", methods=["POST"])
@login_required
def invoice_restock_upload_uzum():
    if not openpyxl:
        return _json_response({"error": "openpyxl library not installed. Please run: pip install openpyxl"}, 500)

    payload = request.get_json(force=True, silent=True) or {}
    items = payload.get("items") or []
    shop_db_id = payload.get("shop_id")

    if not shop_db_id:
        return _json_response({"error": "Shop ID is required"}, 400)

    with SessionLocal() as db:
        shop = db.get(Shop, int(shop_db_id))
        if not shop:
            return _json_response({"error": "Shop not found in DB"}, 404)
        uzum_shop_id = shop.uzum_id

    data_rows = []
    items.sort(key=lambda x: str(x.get("sku") or "").strip().lower())

    for item in items:
        bc = str(item.get("barcode") or "").strip()
        try:
            price = float(item.get("price") or 0)
            qty = int(item.get("qty") or 0)
        except (ValueError, TypeError):
            continue
        if qty > 0:
            data_rows.append([bc, price, qty])

    if not data_rows:
        return _json_response({"error": "No valid items to upload"}, 400)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "\u0422\u043e\u0432\u0430\u0440\u044b \u043d\u0430 \u043e\u0442\u043f\u0440\u0430\u0432\u043a\u0443"
    ws.append(["\u0428\u0442\u0440\u0438\u0445\u043a\u043e\u0434 \u0442\u043e\u0432\u0430\u0440\u0430*", "\u0421\u0435\u0431\u0435\u0441\u0442\u043e\u0438\u043c\u043e\u0441\u0442\u044c (\u0441\u0443\u043c)*", "\u041a\u043e\u043b\u0438\u0447\u0435\u0441\u0442\u0432\u043e (\u0448\u0442)*"])
    for cell in ws[1]: cell.font = Font(bold=True)
    for r in data_rows:
        ws.append(r)

    out = io.BytesIO()
    wb.save(out)
    file_bytes = out.getvalue()

    url = f"https://api-seller.uzum.uz/api/seller/shop/{uzum_shop_id}/v2/invoice/create-from-file"

    # Use current user's key if set, otherwise fall back to admin token.
    api_key = _get_fresh_api_key() or _get_admin_token()
    if not api_key:
        return _json_response({
            "error": "Uzum \u0442\u043e\u043a\u0435\u043d \u043e\u0442\u0441\u0443\u0442\u0441\u0442\u0432\u0443\u0435\u0442. \u0423\u0441\u0442\u0430\u043d\u043e\u0432\u0438\u0442\u0435 Chrome-\u0440\u0430\u0441\u0448\u0438\u0440\u0435\u043d\u0438\u0435 \u00abUzum Token Sync\u00bb "
                     "\u0438\u043b\u0438 \u0432\u0441\u0442\u0430\u0432\u044c\u0442\u0435 \u0442\u043e\u043a\u0435\u043d \u0432\u0440\u0443\u0447\u043d\u0443\u044e \u0432 \u041d\u0430\u0441\u0442\u0440\u043e\u0439\u043a\u0430\u0445."
        }, 401)

    # Warn if already expired before even trying
    exp = _jwt_expires_in_seconds(api_key)
    if exp is not None and exp <= 0:
        return _json_response({
            "error": "Uzum \u0442\u043e\u043a\u0435\u043d \u0438\u0441\u0442\u0451\u043a. \u041e\u0442\u043a\u0440\u043e\u0439\u0442\u0435 \u043a\u0430\u0431\u0438\u043d\u0435\u0442 \u043f\u0440\u043e\u0434\u0430\u0432\u0446\u0430 Uzum \u2014 "
                     "\u0440\u0430\u0441\u0448\u0438\u0440\u0435\u043d\u0438\u0435 \u043e\u0431\u043d\u043e\u0432\u0438\u0442 \u0442\u043e\u043a\u0435\u043d \u0430\u0432\u0442\u043e\u043c\u0430\u0442\u0438\u0447\u0435\u0441\u043a\u0438."
        }, 401)

    headers = {"Authorization": f"Bearer {api_key}" if not api_key.startswith("Bearer ") else api_key}

    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    file_name = f"invoice_restock_{timestamp}.xlsx"

    try:
        res = http_post_multipart(url, file_name, file_bytes, headers)
        return _json_response({"ok": True, "uzum_response": res, "rows_sent": len(data_rows)})
    except Exception as e:
        error_msg = str(e)
        if "401" in error_msg:
            return _json_response({"error": "\u0422\u043e\u043a\u0435\u043d Uzum \u0438\u0441\u0442\u0451\u043a \u0438\u043b\u0438 \u043d\u0435\u0434\u0435\u0439\u0441\u0442\u0432\u0438\u0442\u0435\u043b\u0435\u043d. "
                                   "\u041e\u0442\u043a\u0440\u043e\u0439\u0442\u0435 \u043a\u0430\u0431\u0438\u043d\u0435\u0442 \u043f\u0440\u043e\u0434\u0430\u0432\u0446\u0430 Uzum \u2014 \u0440\u0430\u0441\u0448\u0438\u0440\u0435\u043d\u0438\u0435 \u00abUzum Token Sync\u00bb "
                                   "\u043e\u0431\u043d\u043e\u0432\u0438\u0442 \u0435\u0433\u043e \u0430\u0432\u0442\u043e\u043c\u0430\u0442\u0438\u0447\u0435\u0441\u043a\u0438. \u0415\u0441\u043b\u0438 \u0440\u0430\u0441\u0448\u0438\u0440\u0435\u043d\u0438\u0435 \u043d\u0435 \u0443\u0441\u0442\u0430\u043d\u043e\u0432\u043b\u0435\u043d\u043e, "
                                   "\u0441\u043a\u043e\u043f\u0438\u0440\u0443\u0439\u0442\u0435 \u0441\u0432\u0435\u0436\u0438\u0439 Authorization-\u0442\u043e\u043a\u0435\u043d \u0447\u0435\u0440\u0435\u0437 F12 \u2192 Network \u0438 "
                                   "\u0432\u0441\u0442\u0430\u0432\u044c\u0442\u0435 \u0435\u0433\u043e \u0432 \u041d\u0430\u0441\u0442\u0440\u043e\u0439\u043a\u0430\u0445."}, 401)
        if "403" in error_msg:
            return _json_response({"error": "403 Forbidden: Token is valid but doesn't have permission for this Shop ID."}, 403)
        return _json_response({"error": f"Failed to upload: {error_msg}"}, 500)
