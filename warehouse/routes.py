"""Warehouse routes extracted from app.py as a Flask Blueprint."""
from __future__ import annotations

import io
from datetime import date, datetime, timedelta

from flask import Blueprint, flash, redirect, render_template, request, send_file, url_for
from flask_login import current_user, login_required
from sqlalchemy import select, func, desc

from extensions import SessionLocal, DB_URL_DISPLAY
from models import ProductGroup, Variant, VariantSale
from core.auth_helpers import (
    _current_user_is_admin,
    _json_response,
    _user_shop_ids,
)

try:
    import openpyxl
    from openpyxl.styles import Font
except ImportError:
    openpyxl = None

warehouse_bp = Blueprint("warehouse_bp", __name__)


# ----------------------------
# Warehouse Import/Export helpers
# ----------------------------

WAREHOUSE_EXPORT_HEADERS = ["SKU", "Barcode", "Name", "Warehouse Quantity"]

_WAREHOUSE_HEADER_ALIASES = {
    "sku": {"sku", "article", "artikul"},
    "barcode": {"barcode", "bar code", "ean", "ean13", "штрихкод", "shtrixkod"},
    "quantity": {
        "warehouse quantity", "warehouse qty", "warehouse stock", "warehouse_quantity",
        "qty", "quantity", "stock", "остаток", "склад", "количество", "soni"
    },
}


def _normalize_excel_header(value) -> str:
    text = str(value or "").strip().lower()
    for old, new in (("_", " "), ("-", " "), ("\n", " "), ("\r", " "), ("\t", " ")):
        text = text.replace(old, new)
    return " ".join(text.split())


def _coerce_warehouse_import_qty(value) -> int | None:
    if value is None:
        return None
    if isinstance(value, bool):
        return int(value)
    if isinstance(value, (int, float)):
        return int(float(value))
    text = str(value).strip()
    if not text:
        return None
    try:
        return int(float(text.replace(",", ".")))
    except Exception:
        return None


def _detect_warehouse_import_columns(header_row: tuple) -> dict[str, int]:
    normalized = [_normalize_excel_header(cell) for cell in header_row]
    indices: dict[str, int] = {}
    for field, aliases in _WAREHOUSE_HEADER_ALIASES.items():
        for idx, header in enumerate(normalized):
            if header in aliases:
                indices[field] = idx
                break

    if "sku" not in indices:
        raise ValueError("Column 'SKU' not found in the Excel file.")
    if "quantity" not in indices:
        raise ValueError("Column 'Warehouse Quantity' not found in the Excel file.")
    return indices


def _warehouse_scope_stmt(uid: int):
    allowed_shop_ids = _user_shop_ids(uid)
    stmt = select(Variant, ProductGroup).join(ProductGroup, Variant.group_id == ProductGroup.id)
    if allowed_shop_ids:
        stmt = stmt.where(ProductGroup.shop_id.in_(allowed_shop_ids))
    else:
        stmt = stmt.where(False)
    return stmt, allowed_shop_ids


def _warehouse_page_summary(uid: int) -> dict[str, int]:
    allowed_shop_ids = _user_shop_ids(uid)
    with SessionLocal() as db:
        stmt = (
            select(
                func.count(Variant.id),
                func.count(func.distinct(ProductGroup.shop_id)),
                func.coalesce(func.sum(Variant.warehouse_quantity), 0),
            )
            .join(ProductGroup, Variant.group_id == ProductGroup.id)
        )
        if allowed_shop_ids:
            stmt = stmt.where(ProductGroup.shop_id.in_(allowed_shop_ids))
        else:
            stmt = stmt.where(False)
        variant_count, shop_count, total_qty = db.execute(stmt).one()
        return {
            "variant_count": int(variant_count or 0),
            "shop_count": int(shop_count or 0),
            "total_qty": int(total_qty or 0),
        }


# ----------------------------
# Routes
# ----------------------------

@warehouse_bp.get("/api/health")
def health():
    return _json_response({
        "ok": True,
        "db": "postgresql",
        "db_url": DB_URL_DISPLAY,
    })


@warehouse_bp.get("/api/products")
@login_required
def get_products():
    q = (request.args.get("q") or "").strip()
    days = int(request.args.get("days") or 30)
    since = date.today() - timedelta(days=days)

    with SessionLocal() as db:
        # Return synced variant data for the current product catalog.
        sales_subq = (
            select(VariantSale.variant_id, func.coalesce(func.sum(VariantSale.qty_sold), 0).label("sales_sum"))
            .where(VariantSale.date >= since)
            .group_by(VariantSale.variant_id)
            .subquery()
        )

        stmt = (
            select(
                Variant,
                ProductGroup,
                func.coalesce(sales_subq.c.sales_sum, 0).label("sales_sum")
            )
            .join(ProductGroup, Variant.group_id == ProductGroup.id)
            .outerjoin(sales_subq, Variant.id == sales_subq.c.variant_id)
        )

        if q:
            like = f"%{q}%"
            stmt = stmt.where(
                ProductGroup.name.ilike(like) |
                Variant.sku.ilike(like) |
                Variant.barcode.ilike(like)
            )

        stmt = stmt.order_by(Variant.id)
        rows = db.execute(stmt).all()

        items = []
        for v, g, s_sum in rows:
            name = g.name
            attrs = []
            if v.color: attrs.append(v.color)
            if v.size: attrs.append(v.size)
            if attrs:
                name += f" ({', '.join(attrs)})"

            items.append({
                "id": v.id,
                "name": name,
                "sku": v.sku,
                "barcode": v.barcode,
                "quantity": v.warehouse_quantity,
                "image_url": v.image_url or g.image_url,
                "last30_sales": int(s_sum),
                "created_at": v.created_at.isoformat(),
                "updated_at": v.updated_at.isoformat(),
            })

        return _json_response({
            "days": days,
            "items": items
        })


@warehouse_bp.get("/api/products/<int:variant_id>")
@login_required
def get_product_detail_api(variant_id: int):
    with SessionLocal() as db:
        row = db.execute(
            select(Variant, ProductGroup)
            .join(ProductGroup, Variant.group_id == ProductGroup.id)
            .where(Variant.id == variant_id)
        ).first()

        if not row:
            return _json_response({"error": "Product not found"}, 404)

        v, g = row
        name = g.name
        attrs = []
        if v.color: attrs.append(v.color)
        if v.size: attrs.append(v.size)
        if attrs:
            name += f" ({', '.join(attrs)})"

        return _json_response({
            "id": v.id,
            "name": name,
            "sku": v.sku,
            "barcode": v.barcode,
            "quantity": v.warehouse_quantity,
            "image_url": v.image_url or g.image_url,
            "created_at": v.created_at.isoformat(),
            "updated_at": v.updated_at.isoformat(),
        })


@warehouse_bp.route("/warehouse/data", methods=["GET"])
@login_required
def warehouse_data_page():
    return redirect(url_for("warehouse_bp.warehouse_import"), code=302)


@warehouse_bp.route("/warehouse/import", methods=["GET", "POST"])
@login_required
def warehouse_import():
    if not openpyxl:
        return "openpyxl library not installed", 500

    uid = int(current_user.get_id())

    if request.method == "GET":
        summary = _warehouse_page_summary(uid)
        return render_template(
            "warehouse_data.html",
            title="Импорт склада Excel",
            summary=summary,
            expected_headers=WAREHOUSE_EXPORT_HEADERS,
        )

    file = request.files.get("file")
    if not file or not file.filename:
        flash("Выберите Excel файл .xlsx для импорта.")
        return redirect(url_for("warehouse_bp.warehouse_import"))

    if not file.filename.lower().endswith(".xlsx"):
        flash("Поддерживаются только файлы Excel формата .xlsx.")
        return redirect(url_for("warehouse_bp.warehouse_import"))

    try:
        wb = openpyxl.load_workbook(file, data_only=True)
        ws = wb[wb.sheetnames[0]]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            raise ValueError("Excel file is empty.")

        column_map = _detect_warehouse_import_columns(rows[0])
        idx_sku = column_map["sku"]
        idx_qty = column_map["quantity"]
        idx_barcode = column_map.get("barcode")

        stmt, allowed_shop_ids = _warehouse_scope_stmt(uid)
        with SessionLocal() as db:
            scoped_variants = db.execute(stmt).all()
            sku_map = {
                (variant.sku or "").strip().lower(): variant
                for variant, _group in scoped_variants
                if variant.sku
            }
            barcode_map = {
                (variant.barcode or "").strip().lower(): variant
                for variant, _group in scoped_variants
                if variant.barcode
            }

            matched_count = 0
            changed_count = 0
            barcode_match_count = 0
            missing_count = 0
            skipped_count = 0
            missing_examples: list[str] = []

            for excel_row_number, row in enumerate(rows[1:], start=2):
                if not row or all(cell in (None, "") for cell in row):
                    continue

                sku = str(row[idx_sku]).strip() if len(row) > idx_sku and row[idx_sku] else ""
                barcode = ""
                if idx_barcode is not None and len(row) > idx_barcode and row[idx_barcode]:
                    barcode = str(row[idx_barcode]).strip()

                qty_val = row[idx_qty] if len(row) > idx_qty else None
                qty = _coerce_warehouse_import_qty(qty_val)
                if qty is None:
                    skipped_count += 1
                    continue

                variant = None
                used_barcode = False
                if sku:
                    variant = sku_map.get(sku.lower())
                if not variant and barcode:
                    variant = barcode_map.get(barcode.lower())
                    used_barcode = variant is not None

                if not variant:
                    missing_count += 1
                    label = sku or barcode or f"row {excel_row_number}"
                    if len(missing_examples) < 5:
                        missing_examples.append(label)
                    continue

                matched_count += 1
                if used_barcode:
                    barcode_match_count += 1
                if (variant.warehouse_quantity or 0) != qty:
                    variant.warehouse_quantity = qty
                    changed_count += 1

            db.commit()

        flash_parts = [
            f"Импорт завершён: найдено {matched_count}, изменено {changed_count}, пропущено {skipped_count}."
        ]
        if barcode_match_count:
            flash_parts.append(f"По штрихкоду сопоставлено {barcode_match_count}.")
        if missing_count:
            msg = f"Не найдено {missing_count}"
            if missing_examples:
                msg += f" ({', '.join(missing_examples)})"
            flash_parts.append(msg + ".")
        flash(" ".join(flash_parts))
        return redirect(url_for("warehouse_bp.warehouse_import"))

    except Exception as e:
        flash(f"Ошибка импорта: {str(e)}")
        return redirect(url_for("warehouse_bp.warehouse_import"))


@warehouse_bp.route("/warehouse/export", methods=["GET"])
@login_required
def warehouse_export():
    if not openpyxl:
        return "openpyxl library not installed", 500

    uid = int(current_user.get_id())
    stmt, allowed_shop_ids = _warehouse_scope_stmt(uid)
    with SessionLocal() as db:
        stmt = stmt.order_by(ProductGroup.name, Variant.sku)
        rows = db.execute(stmt).all()

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Warehouse Data"

        ws.append(WAREHOUSE_EXPORT_HEADERS)
        for cell in ws[1]:
            cell.font = Font(bold=True)

        for v, g in rows:
            name = g.name
            if v.color or v.size:
                attrs = [x for x in [v.color, v.size] if x]
                name += f" ({', '.join(attrs)})"

            ws.append([
                v.sku,
                v.barcode,
                name,
                v.warehouse_quantity,
            ])

        out = io.BytesIO()
        wb.save(out)
        out.seek(0)

        filename = f"warehouse_qty_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
        return send_file(
            out,
            download_name=filename,
            as_attachment=True,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )


@warehouse_bp.get("/api/summary")
@login_required
def summary():
    days = int(request.args.get("days") or 30)
    since = date.today() - timedelta(days=days)

    with SessionLocal() as db:
        # Return top selling variants for the summary table
        sales_subq = (
            select(VariantSale.variant_id, func.coalesce(func.sum(VariantSale.qty_sold), 0).label("sales_sum"))
            .where(VariantSale.date >= since)
            .group_by(VariantSale.variant_id)
            .subquery()
        )

        uid = int(current_user.get_id())
        allowed_shop_ids = _user_shop_ids(uid)
        stmt = (
            select(Variant, ProductGroup, sales_subq.c.sales_sum)
            .join(ProductGroup, Variant.group_id == ProductGroup.id)
            .join(sales_subq, Variant.id == sales_subq.c.variant_id)
        )
        if allowed_shop_ids:
            stmt = stmt.where(ProductGroup.shop_id.in_(allowed_shop_ids))
        else:
            stmt = stmt.where(False)
        stmt = stmt.order_by(desc(sales_subq.c.sales_sum)).limit(50)
        rows = db.execute(stmt).all()

        items = []
        for v, g, s_sum in rows:
            items.append({
                "id": v.id,
                "name": g.name,
                "sku": v.sku,
                "barcode": v.barcode,
                "quantity": v.warehouse_quantity,
                "image_url": v.image_url or g.image_url,
                "last30_sales": int(s_sum),
            })

    return _json_response({
        "days": days,
        "items": items
    })
